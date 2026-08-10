"""Server-generated Insurance Metadata Workbook V2 contract."""

from __future__ import annotations

from datetime import date, datetime
import hashlib
from io import BytesIO
import json
from pathlib import PurePosixPath
from typing import Any, Annotated, Literal, TypeAlias, TypeVar, cast
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import (  # type: ignore[import-untyped]
    DataValidation,
)
from pydantic import ConfigDict, Field, StrictBytes, StrictStr, StringConstraints

from proof_agent.capabilities.knowledge.hybrid.metadata_review import (
    InsuranceMetadataProfileRevision,
    InsuranceMetadataReviewDecisionV2,
    InsuranceMetadataReviewSet,
    InsuranceMetadataReviewV2,
    MetadataReviewConflictError,
    advance_insurance_metadata_review_set,
    create_insurance_metadata_override,
    save_metadata_review_draft,
)
from proof_agent.contracts._base import FrozenModel
from proof_agent.contracts.insurance_rules import InsuranceRuleMetadataDraft
from proof_agent.contracts.knowledge_index import ExactArtifactRef


NonBlankStr = Annotated[StrictStr, StringConstraints(strip_whitespace=True, min_length=1)]
Sha256 = Annotated[StrictStr, StringConstraints(pattern=r"^[0-9a-f]{64}$")]

TEMPLATE_REVISION: Literal["insurance-rule-metadata.v2"] = (
    "insurance-rule-metadata.v2"
)
SHEET_NAMES = (
    "Instructions",
    "Document Defaults",
    "Rule Unit Overrides",
    "Reference Values",
    "_Manifest",
)
GovernedField: TypeAlias = Literal[
    "authority",
    "effective_from",
    "effective_to",
    "taxonomy_id",
    "taxonomy_revision_id",
    "precedence_policy_revision_id",
    "precedence_authority_tier",
    "precedence_order",
]
WorkbookCellValue: TypeAlias = str | int | date | None
MergeClassification: TypeAlias = Literal[
    "unchanged",
    "workbook_only",
    "server_only",
    "matching_change",
    "conflict",
]
OverrideMode: TypeAlias = Literal["inherit", "override"]
_MergeValue = TypeVar("_MergeValue")

GOVERNED_FIELDS: tuple[GovernedField, ...] = (
    "authority",
    "effective_from",
    "effective_to",
    "taxonomy_id",
    "taxonomy_revision_id",
    "precedence_policy_revision_id",
    "precedence_authority_tier",
    "precedence_order",
)


class _WorkbookV2Model(FrozenModel):
    model_config = ConfigDict(frozen=True, extra="forbid", strict=True)


class WorkbookRuleUnitInventoryItem(_WorkbookV2Model):
    canonical_anchor: NonBlankStr
    citation_uri: NonBlankStr
    safe_preview: Annotated[StrictStr, StringConstraints(max_length=512)]


class WorkbookMetadataValuesV2(_WorkbookV2Model):
    authority: NonBlankStr | None = None
    effective_from: date | None = None
    effective_to: date | None = None
    taxonomy_id: NonBlankStr | None = None
    taxonomy_revision_id: NonBlankStr | None = None
    precedence_policy_revision_id: NonBlankStr | None = None
    precedence_authority_tier: NonBlankStr | None = None
    precedence_order: int | None = Field(default=None, ge=0, strict=True)


class WorkbookReviewBaseV2(_WorkbookV2Model):
    scope: Literal["document_default", "rule_unit_override"]
    canonical_anchor: NonBlankStr | None = None
    override_mode: Literal["default", "inherit", "override"]
    review_id: NonBlankStr | None = None
    review_identity: Sha256 | None = None
    review_version: int | None = Field(default=None, ge=1, strict=True)
    values: WorkbookMetadataValuesV2


class MetadataWorkbookExportManifestV2(_WorkbookV2Model):
    schema_version: Literal["insurance-metadata-workbook-export-manifest.v2"] = (
        "insurance-metadata-workbook-export-manifest.v2"
    )
    template_revision: Literal["insurance-rule-metadata.v2"] = TEMPLATE_REVISION
    export_id: NonBlankStr
    environment_id: NonBlankStr
    source_id: NonBlankStr
    document_id: NonBlankStr
    revision_id: NonBlankStr
    structured_build_id: NonBlankStr
    profile_revision_id: NonBlankStr
    review_set_id: NonBlankStr
    review_set_identity: Sha256
    review_set_generation: int = Field(ge=1, strict=True)
    exported_at: datetime
    expires_at: datetime
    canonical_anchor_set_sha256: Sha256
    canonical_anchors: tuple[NonBlankStr, ...] = Field(max_length=10_000)
    registered_validation_ranges: tuple[NonBlankStr, ...]
    base_reviews: tuple[WorkbookReviewBaseV2, ...] = Field(
        min_length=1, max_length=10_001
    )


class MetadataWorkbookExportV2(_WorkbookV2Model):
    manifest: MetadataWorkbookExportManifestV2
    content: StrictBytes


class MetadataWorkbookValidationError(ValueError):
    """Returned Workbook V2 failed a controlled, content-safe check."""


class WorkbookFieldMergeV2(_WorkbookV2Model):
    scope: Literal["document_default", "rule_unit_override"]
    canonical_anchor: NonBlankStr | None = None
    field: Literal[
        "authority",
        "effective_from",
        "effective_to",
        "taxonomy_id",
        "taxonomy_revision_id",
        "precedence_policy_revision_id",
        "precedence_authority_tier",
        "precedence_order",
    ]
    base_value: str | int | date | None = None
    server_value: str | int | date | None = None
    workbook_value: str | int | date | None = None
    classification: Literal[
        "unchanged",
        "workbook_only",
        "server_only",
        "matching_change",
        "conflict",
    ]
    proposed_value: str | int | date | None = None


class WorkbookOverrideModeMergeV2(_WorkbookV2Model):
    canonical_anchor: NonBlankStr
    base_mode: Literal["inherit", "override"]
    server_mode: Literal["inherit", "override"]
    workbook_mode: Literal["inherit", "override"]
    classification: Literal[
        "unchanged",
        "workbook_only",
        "server_only",
        "matching_change",
        "conflict",
    ]
    proposed_mode: Literal["inherit", "override"] | None = None
    override_reason: Annotated[StrictStr, StringConstraints(max_length=2_000)] | None = None


class MetadataWorkbookImportPreviewV2(_WorkbookV2Model):
    schema_version: Literal["insurance-metadata-workbook-import-preview.v2"] = (
        "insurance-metadata-workbook-import-preview.v2"
    )
    preview_id: NonBlankStr
    preview_identity: Sha256
    export_id: NonBlankStr
    source_id: NonBlankStr
    document_id: NonBlankStr
    revision_id: NonBlankStr
    profile_revision_id: NonBlankStr
    export_review_set_identity: Sha256
    current_review_set_identity: Sha256
    current_review_set_generation: int = Field(ge=1, strict=True)
    previewed_at: datetime
    state: Literal["ready_to_apply", "conflicts"]
    conflict_count: int = Field(ge=0, strict=True)
    field_merges: tuple[WorkbookFieldMergeV2, ...] = Field(
        min_length=len(GOVERNED_FIELDS), max_length=80_008
    )
    override_modes: tuple[WorkbookOverrideModeMergeV2, ...] = Field(
        max_length=10_000
    )


class MetadataWorkbookApplyResultV2(_WorkbookV2Model):
    review_set: InsuranceMetadataReviewSet
    decisions: tuple[InsuranceMetadataReviewDecisionV2, ...]


class MetadataWorkbookExportAuthorityV2(_WorkbookV2Model):
    schema_version: Literal["insurance-metadata-workbook-export-authority.v2"] = (
        "insurance-metadata-workbook-export-authority.v2"
    )
    manifest: MetadataWorkbookExportManifestV2
    artifact_ref: ExactArtifactRef
    state: Literal["available", "consumed", "expired", "stale"]
    created_by: NonBlankStr
    downloaded_at: datetime | None = None
    consumed_at: datetime | None = None


class MetadataWorkbookValidationIssueV2(_WorkbookV2Model):
    sheet: Annotated[StrictStr, StringConstraints(max_length=255)] | None = None
    row: int | None = Field(default=None, ge=1, strict=True)
    field: Annotated[StrictStr, StringConstraints(max_length=255)] | None = None
    code: NonBlankStr
    suggested_action_key: NonBlankStr


class MetadataWorkbookValidationReportV2(_WorkbookV2Model):
    schema_version: Literal["insurance-metadata-workbook-validation-report.v2"] = (
        "insurance-metadata-workbook-validation-report.v2"
    )
    total_error_count: int = Field(ge=1, strict=True)
    errors: tuple[MetadataWorkbookValidationIssueV2, ...] = Field(
        min_length=1,
        max_length=100,
    )


class MetadataWorkbookImportPreviewAuthorityV2(_WorkbookV2Model):
    schema_version: Literal[
        "insurance-metadata-workbook-import-preview-authority.v2"
    ] = "insurance-metadata-workbook-import-preview-authority.v2"
    preview_id: NonBlankStr
    source_id: NonBlankStr
    export_id: NonBlankStr
    original_ref: ExactArtifactRef
    state: Literal[
        "validation_failed",
        "conflicts",
        "ready_to_apply",
        "applied",
        "expired",
        "stale",
    ]
    preview: MetadataWorkbookImportPreviewV2 | None = None
    validation_report: MetadataWorkbookValidationReportV2 | None = None
    created_by: NonBlankStr
    created_at: datetime
    expires_at: datetime
    applied_at: datetime | None = None


class MetadataWorkbookApplyCommitV2(_WorkbookV2Model):
    source_revision: int = Field(ge=1, strict=True)
    review_set: InsuranceMetadataReviewSet
    decisions: tuple[InsuranceMetadataReviewDecisionV2, ...]
    preview: MetadataWorkbookImportPreviewAuthorityV2


def generate_metadata_workbook_v2(
    *,
    export_id: str,
    environment_id: str,
    review_set: InsuranceMetadataReviewSet,
    profile: InsuranceMetadataProfileRevision,
    rule_units: tuple[WorkbookRuleUnitInventoryItem, ...],
    exported_at: datetime,
    expires_at: datetime,
) -> MetadataWorkbookExportV2:
    """Freeze a Profile-bound Review Set into one governed XLSX export."""

    if review_set.profile_revision_id != profile.profile_revision_id:
        raise ValueError("Workbook Profile does not match the Review Set binding")
    if exported_at.tzinfo is None or exported_at.utcoffset() is None:
        raise ValueError("Workbook export timestamp must be timezone-aware")
    if expires_at.tzinfo is None or expires_at.utcoffset() is None:
        raise ValueError("Workbook expiry timestamp must be timezone-aware")
    if expires_at <= exported_at:
        raise ValueError("Workbook expiry must follow export time")
    if len(rule_units) > 10_000:
        raise ValueError("metadata_review_capacity_exceeded")
    anchors = tuple(item.canonical_anchor for item in rule_units)
    if len(anchors) != len(set(anchors)):
        raise ValueError("Workbook Rule Unit inventory requires unique anchors")
    override_by_anchor = {
        review.canonical_anchor: review
        for review in review_set.reviews
        if review.scope == "rule_unit_override"
    }
    if set(override_by_anchor) - set(anchors):
        raise ValueError("Workbook inventory does not cover every Review override")
    default_review = review_set.reviews[0]
    range_rows = _reference_value_rows(profile)
    registered_ranges = tuple(
        f"'Reference Values'!$B${start}:$B${end}"
        for _kind, start, end in range_rows.ranges
    )
    bases = [
        _base_review(
            default_review,
            override_mode="default",
        )
    ]
    for item in rule_units:
        review = override_by_anchor.get(item.canonical_anchor)
        bases.append(
            WorkbookReviewBaseV2(
                scope="rule_unit_override",
                canonical_anchor=item.canonical_anchor,
                override_mode="inherit" if review is None else "override",
                review_id=None if review is None else review.review_id,
                review_identity=None if review is None else review.review_identity,
                review_version=None if review is None else review.review_version,
                values=_metadata_values(
                    default_review.current_draft
                    if review is None
                    else review.current_draft
                ),
            )
        )
    manifest = MetadataWorkbookExportManifestV2(
        export_id=export_id,
        environment_id=environment_id,
        source_id=review_set.source_id,
        document_id=review_set.document_id,
        revision_id=review_set.revision_id,
        structured_build_id=review_set.structured_build_id,
        profile_revision_id=review_set.profile_revision_id,
        review_set_id=review_set.review_set_id,
        review_set_identity=review_set.review_set_identity,
        review_set_generation=review_set.generation,
        exported_at=exported_at,
        expires_at=expires_at,
        canonical_anchor_set_sha256=_sha256(anchors),
        canonical_anchors=anchors,
        registered_validation_ranges=registered_ranges,
        base_reviews=tuple(bases),
    )
    content = _build_workbook(
        manifest=manifest,
        profile=profile,
        review_set=review_set,
        rule_units=rule_units,
        reference_rows=range_rows,
    )
    return MetadataWorkbookExportV2(manifest=manifest, content=content)


def create_metadata_workbook_import_preview_v2(
    *,
    preview_id: str,
    export_manifest: MetadataWorkbookExportManifestV2,
    returned_content: bytes,
    current_review_set: InsuranceMetadataReviewSet,
    profile: InsuranceMetadataProfileRevision,
    previewed_at: datetime,
) -> MetadataWorkbookImportPreviewV2:
    """Validate one returned export and compute a non-mutating three-way merge."""

    _validate_preview_authority(
        export_manifest=export_manifest,
        current_review_set=current_review_set,
        profile=profile,
        previewed_at=previewed_at,
    )
    _preflight_xlsx(returned_content)
    try:
        workbook = load_workbook(
            BytesIO(returned_content),
            data_only=False,
            read_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise MetadataWorkbookValidationError(
            "metadata_workbook_package_invalid"
        ) from exc
    if tuple(workbook.sheetnames) != SHEET_NAMES:
        raise MetadataWorkbookValidationError("metadata_workbook_sheet_set_invalid")
    if workbook["_Manifest"].sheet_state != "hidden":
        raise MetadataWorkbookValidationError("metadata_workbook_manifest_visibility_invalid")
    _reject_formulas_and_oversized_cells(workbook)
    _validate_returned_manifest(workbook["_Manifest"], export_manifest)
    _validate_validation_ranges(workbook, export_manifest)

    returned, returned_modes = _returned_review_values(
        workbook, export_manifest, profile
    )
    current_default = current_review_set.reviews[0]
    current_by_anchor = {
        review.canonical_anchor: review
        for review in current_review_set.reviews
        if review.scope == "rule_unit_override"
    }
    bases_by_anchor = {
        base.canonical_anchor: base
        for base in export_manifest.base_reviews
        if base.scope == "rule_unit_override"
    }
    merges: list[WorkbookFieldMergeV2] = []
    mode_merges: list[WorkbookOverrideModeMergeV2] = []
    default_base = export_manifest.base_reviews[0]
    merges.extend(
        _merge_values(
            scope="document_default",
            canonical_anchor=None,
            base=default_base.values,
            server=_metadata_values(current_default.current_draft),
            workbook=returned[("document_default", None)],
        )
    )
    for anchor in export_manifest.canonical_anchors:
        base = bases_by_anchor[anchor]
        if base.override_mode not in {"inherit", "override"}:
            raise MetadataWorkbookValidationError(
                "metadata_workbook_override_base_mode_invalid"
            )
        base_mode = cast(OverrideMode, base.override_mode)
        current = current_by_anchor.get(anchor)
        workbook_mode, override_reason = returned_modes[anchor]
        server_mode: OverrideMode = "inherit" if current is None else "override"
        mode_classification, proposed_mode = _classify_merge(
            base_mode,
            server_mode,
            workbook_mode,
        )
        mode_merges.append(
            WorkbookOverrideModeMergeV2(
                canonical_anchor=anchor,
                base_mode=base_mode,
                server_mode=server_mode,
                workbook_mode=workbook_mode,
                classification=mode_classification,
                proposed_mode=proposed_mode,
                override_reason=override_reason,
            )
        )
        server = _metadata_values(
            current_default.current_draft if current is None else current.current_draft
        )
        merges.extend(
            _merge_values(
                scope="rule_unit_override",
                canonical_anchor=anchor,
                base=base.values,
                server=server,
                workbook=returned[("rule_unit_override", anchor)],
            )
        )
    conflict_count = sum(merge.classification == "conflict" for merge in merges) + sum(
        merge.classification == "conflict" for merge in mode_merges
    )
    material = {
        "preview_id": preview_id,
        "export_id": export_manifest.export_id,
        "current_review_set_identity": current_review_set.review_set_identity,
        "field_merges": [merge.model_dump(mode="json") for merge in merges],
        "override_modes": [merge.model_dump(mode="json") for merge in mode_merges],
    }
    return MetadataWorkbookImportPreviewV2(
        preview_id=preview_id,
        preview_identity=_sha256(material),
        export_id=export_manifest.export_id,
        source_id=export_manifest.source_id,
        document_id=export_manifest.document_id,
        revision_id=export_manifest.revision_id,
        profile_revision_id=export_manifest.profile_revision_id,
        export_review_set_identity=export_manifest.review_set_identity,
        current_review_set_identity=current_review_set.review_set_identity,
        current_review_set_generation=current_review_set.generation,
        previewed_at=previewed_at,
        state="conflicts" if conflict_count else "ready_to_apply",
        conflict_count=conflict_count,
        field_merges=tuple(merges),
        override_modes=tuple(mode_merges),
    )


def apply_metadata_workbook_import_preview_v2(
    preview: MetadataWorkbookImportPreviewV2,
    *,
    current_review_set: InsuranceMetadataReviewSet,
    profile: InsuranceMetadataProfileRevision,
    expected_preview_identity: str,
    actor: str,
    reason: str,
) -> MetadataWorkbookApplyResultV2:
    """Apply one exact conflict-free Preview to an immutable Review Set value."""

    if preview.preview_identity != expected_preview_identity:
        raise MetadataReviewConflictError("metadata Workbook Preview identity changed")
    if preview.state != "ready_to_apply" or preview.conflict_count:
        raise MetadataReviewConflictError(
            "metadata Workbook Preview has unresolved conflicts"
        )
    if (
        current_review_set.review_set_identity
        != preview.current_review_set_identity
        or current_review_set.generation != preview.current_review_set_generation
    ):
        raise MetadataReviewConflictError(
            "metadata Review Set changed after Workbook Preview"
        )
    if (
        current_review_set.source_id != preview.source_id
        or current_review_set.document_id != preview.document_id
        or current_review_set.revision_id != preview.revision_id
        or current_review_set.profile_revision_id != preview.profile_revision_id
        or profile.profile_revision_id != preview.profile_revision_id
    ):
        raise MetadataReviewConflictError(
            "metadata Workbook Preview authority changed"
        )
    if not actor.strip() or not reason.strip():
        raise ValueError("Workbook Apply actor and reason must be nonblank")

    changes_by_review: dict[
        tuple[str, str | None], dict[str, str | int | date | None]
    ] = {}
    for merge in preview.field_merges:
        if (
            merge.classification == "workbook_only"
            and merge.proposed_value != merge.server_value
        ):
            changes_by_review.setdefault(
                (merge.scope, merge.canonical_anchor), {}
            )[merge.field] = merge.proposed_value

    mode_by_anchor = {
        merge.canonical_anchor: merge for merge in preview.override_modes
    }

    updated_set = current_review_set
    decisions: list[InsuranceMetadataReviewDecisionV2] = []
    for (scope, anchor), changes in changes_by_review.items():
        review = next(
            (
                item
                for item in updated_set.reviews
                if item.scope == scope and item.canonical_anchor == anchor
            ),
            None,
        )
        if review is None:
            mode = mode_by_anchor.get(anchor or "")
            if (
                scope != "rule_unit_override"
                or mode is None
                or mode.proposed_mode != "override"
                or mode.classification not in {"workbook_only", "matching_change"}
            ):
                raise MetadataReviewConflictError(
                    "Workbook Apply cannot create an unreviewed override"
                )
            created = create_insurance_metadata_override(
                updated_set,
                canonical_anchor=anchor or "",
                profile=profile,
                actor=actor,
                reason=reason,
                changes=changes,
                action="workbook_apply",
            )
            updated_set = created.review_set
            decisions.append(created.decision)
            continue
        result = save_metadata_review_draft(
            review,
            profile=profile,
            expected_review_version=review.review_version,
            expected_review_identity=review.review_identity,
            actor=actor,
            reason=reason,
            changes=changes,
            action="workbook_apply",
        )
        updated_set = advance_insurance_metadata_review_set(
            updated_set,
            result.review,
        )
        decisions.append(result.decision)
    return MetadataWorkbookApplyResultV2(
        review_set=updated_set,
        decisions=tuple(decisions),
    )


def _validate_preview_authority(
    *,
    export_manifest: MetadataWorkbookExportManifestV2,
    current_review_set: InsuranceMetadataReviewSet,
    profile: InsuranceMetadataProfileRevision,
    previewed_at: datetime,
) -> None:
    if previewed_at.tzinfo is None or previewed_at.utcoffset() is None:
        raise MetadataWorkbookValidationError("metadata_workbook_preview_time_invalid")
    if previewed_at > export_manifest.expires_at:
        raise MetadataWorkbookValidationError("metadata_workbook_export_expired")
    if (
        current_review_set.source_id != export_manifest.source_id
        or current_review_set.document_id != export_manifest.document_id
        or current_review_set.revision_id != export_manifest.revision_id
        or current_review_set.structured_build_id
        != export_manifest.structured_build_id
        or current_review_set.profile_revision_id
        != export_manifest.profile_revision_id
        or profile.profile_revision_id != export_manifest.profile_revision_id
    ):
        raise MetadataWorkbookValidationError("metadata_workbook_structurally_stale")
    current_anchors = {
        review.canonical_anchor
        for review in current_review_set.reviews
        if review.scope == "rule_unit_override"
    }
    if not current_anchors.issubset(set(export_manifest.canonical_anchors)):
        raise MetadataWorkbookValidationError("metadata_workbook_anchor_set_stale")


_DENIED_PACKAGE_MARKERS = (
    "vbaProject.bin",
    "xl/externalLinks/",
    "xl/activeX/",
    "xl/embeddings/",
    "xl/connections.xml",
    "xl/pivotCache/",
)


def _preflight_xlsx(content: bytes) -> None:
    if not content or len(content) > 10 * 1024 * 1024:
        raise MetadataWorkbookValidationError("metadata_workbook_compressed_limit")
    try:
        with ZipFile(BytesIO(content)) as package:
            names = [item.filename for item in package.infolist()]
            if len(names) != len(set(names)) or len(names) != len(set(map(str.casefold, names))):
                raise MetadataWorkbookValidationError(
                    "metadata_workbook_package_member_ambiguous"
                )
            expanded = 0
            for item in package.infolist():
                path = PurePosixPath(item.filename)
                if (
                    path.is_absolute()
                    or ".." in path.parts
                    or "\\" in item.filename
                    or any(marker.casefold() in item.filename.casefold() for marker in _DENIED_PACKAGE_MARKERS)
                ):
                    raise MetadataWorkbookValidationError(
                        "metadata_workbook_package_member_denied"
                    )
                expanded += item.file_size
                if expanded > 80 * 1024 * 1024:
                    raise MetadataWorkbookValidationError(
                        "metadata_workbook_expanded_limit"
                    )
    except BadZipFile as exc:
        raise MetadataWorkbookValidationError("metadata_workbook_package_invalid") from exc


def _reject_formulas_and_oversized_cells(workbook: Any) -> None:
    if tuple(workbook.defined_names) != ():
        raise MetadataWorkbookValidationError("metadata_workbook_defined_name_invalid")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise MetadataWorkbookValidationError(
                        "metadata_workbook_formula_forbidden"
                    )
                if isinstance(cell.value, str) and len(cell.value) > 4_096:
                    raise MetadataWorkbookValidationError(
                        "metadata_workbook_cell_limit"
                    )


def _validate_returned_manifest(
    sheet: Any,
    manifest: MetadataWorkbookExportManifestV2,
) -> None:
    expected = manifest.model_dump(mode="json")
    for row in range(2, 17):
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if key not in expected or str(value) != str(expected[key]):
            raise MetadataWorkbookValidationError(
                "metadata_workbook_manifest_identity_invalid"
            )


def _validate_validation_ranges(
    workbook: Any,
    manifest: MetadataWorkbookExportManifestV2,
) -> None:
    allowed = set(manifest.registered_validation_ranges) | {'"inherit,override"'}
    observed: set[str] = set()
    for name in ("Document Defaults", "Rule Unit Overrides"):
        for validation in workbook[name].data_validations.dataValidation:
            if validation.type != "list" or validation.formula1 not in allowed:
                raise MetadataWorkbookValidationError(
                    "metadata_workbook_validation_range_invalid"
                )
            observed.add(str(validation.formula1))
    if not set(manifest.registered_validation_ranges).issubset(observed):
        raise MetadataWorkbookValidationError(
            "metadata_workbook_validation_range_missing"
        )


def _returned_review_values(
    workbook: Any,
    manifest: MetadataWorkbookExportManifestV2,
    profile: InsuranceMetadataProfileRevision,
) -> tuple[
    dict[tuple[str, str | None], WorkbookMetadataValuesV2],
    dict[str, tuple[Literal["inherit", "override"], str | None]],
]:
    defaults = workbook["Document Defaults"]
    if tuple(defaults.cell(row=5, column=index).value for index in range(1, 18)) != _DEFAULT_HEADERS:
        raise MetadataWorkbookValidationError("metadata_workbook_default_columns_invalid")
    base_default = manifest.base_reviews[0]
    locked = tuple(defaults.cell(row=6, column=index).value for index in range(1, 10))
    expected_locked = (
        base_default.review_id,
        base_default.review_identity,
        base_default.review_version,
        manifest.document_id,
        manifest.revision_id,
        manifest.structured_build_id,
        manifest.profile_revision_id,
        defaults.cell(row=6, column=8).value,
        "document_default",
    )
    if locked != expected_locked:
        raise MetadataWorkbookValidationError("metadata_workbook_default_identity_invalid")
    result: dict[tuple[str, str | None], WorkbookMetadataValuesV2] = {
        ("document_default", None): _values_from_cells(
            tuple(defaults.cell(row=6, column=index).value for index in range(10, 18)),
            profile,
        )
    }
    overrides = workbook["Rule Unit Overrides"]
    if tuple(overrides.cell(row=5, column=index).value for index in range(1, 19)) != _OVERRIDE_HEADERS:
        raise MetadataWorkbookValidationError("metadata_workbook_override_columns_invalid")
    modes: dict[str, tuple[Literal["inherit", "override"], str | None]] = {}
    for offset, anchor in enumerate(manifest.canonical_anchors, 6):
        if overrides.cell(row=offset, column=1).value != anchor:
            raise MetadataWorkbookValidationError("metadata_workbook_anchor_identity_invalid")
        mode = overrides.cell(row=offset, column=9).value
        if mode not in {"inherit", "override"}:
            raise MetadataWorkbookValidationError("metadata_workbook_override_mode_invalid")
        raw_reason = overrides.cell(row=offset, column=10).value
        if raw_reason is not None and (
            not isinstance(raw_reason, str) or len(raw_reason.strip()) > 2_000
        ):
            raise MetadataWorkbookValidationError("metadata_workbook_override_reason_invalid")
        modes[anchor] = (
            mode,
            None if raw_reason is None else raw_reason.strip() or None,
        )
        result[("rule_unit_override", anchor)] = _values_from_cells(
            tuple(overrides.cell(row=offset, column=index).value for index in range(11, 19)),
            profile,
        )
    return result, modes


def _values_from_cells(
    raw: tuple[object, ...],
    profile: InsuranceMetadataProfileRevision,
) -> WorkbookMetadataValuesV2:
    normalized = tuple(_cell_value(field, value) for field, value in zip(GOVERNED_FIELDS, raw, strict=True))
    values = WorkbookMetadataValuesV2.model_validate(
        dict(zip(GOVERNED_FIELDS, normalized, strict=True))
    )
    if values.authority is not None and values.authority not in profile.authority_codes:
        raise MetadataWorkbookValidationError("metadata_workbook_authority_invalid")
    if values.taxonomy_id not in {None, profile.taxonomy_id}:
        raise MetadataWorkbookValidationError("metadata_workbook_taxonomy_invalid")
    if values.taxonomy_revision_id not in {None, profile.taxonomy_revision_id}:
        raise MetadataWorkbookValidationError("metadata_workbook_taxonomy_revision_invalid")
    if values.precedence_policy_revision_id not in {
        None,
        profile.precedence_policy_revision_id,
    }:
        raise MetadataWorkbookValidationError("metadata_workbook_precedence_policy_invalid")
    if (
        values.precedence_authority_tier is not None
        and values.precedence_authority_tier not in profile.precedence_authority_tiers
    ):
        raise MetadataWorkbookValidationError("metadata_workbook_precedence_tier_invalid")
    if (
        values.effective_from is not None
        and values.effective_to is not None
        and values.effective_to < values.effective_from
    ):
        raise MetadataWorkbookValidationError("metadata_workbook_effective_range_invalid")
    return values


def _cell_value(field: str, value: object) -> str | int | date | None:
    if value is None:
        return None
    if field in {"effective_from", "effective_to"}:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raise MetadataWorkbookValidationError("metadata_workbook_date_invalid")
    if field == "precedence_order":
        if isinstance(value, bool) or not isinstance(value, int) or value < 0:
            raise MetadataWorkbookValidationError("metadata_workbook_order_invalid")
        return value
    if not isinstance(value, str) or not value.strip():
        raise MetadataWorkbookValidationError("metadata_workbook_text_invalid")
    return value.strip()


def _merge_values(
    *,
    scope: Literal["document_default", "rule_unit_override"],
    canonical_anchor: str | None,
    base: WorkbookMetadataValuesV2,
    server: WorkbookMetadataValuesV2,
    workbook: WorkbookMetadataValuesV2,
) -> list[WorkbookFieldMergeV2]:
    merges: list[WorkbookFieldMergeV2] = []
    for field in GOVERNED_FIELDS:
        base_value = cast(WorkbookCellValue, getattr(base, field))
        server_value = cast(WorkbookCellValue, getattr(server, field))
        workbook_value = cast(WorkbookCellValue, getattr(workbook, field))
        classification, proposed = _classify_merge(
            base_value,
            server_value,
            workbook_value,
        )
        merges.append(
            WorkbookFieldMergeV2(
                scope=scope,
                canonical_anchor=canonical_anchor,
                field=field,
                base_value=base_value,
                server_value=server_value,
                workbook_value=workbook_value,
                classification=classification,
                proposed_value=proposed,
            )
        )
    return merges


def _classify_merge(
    base: _MergeValue,
    server: _MergeValue,
    workbook: _MergeValue,
) -> tuple[MergeClassification, _MergeValue | None]:
    if server == base and workbook == base:
        return "unchanged", server
    if server == base:
        return "workbook_only", workbook
    if workbook == base:
        return "server_only", server
    if server == workbook:
        return "matching_change", server
    return "conflict", None


class _ReferenceRows:
    def __init__(
        self,
        *,
        values: list[tuple[str, str, str, str | None]],
        ranges: list[tuple[str, int, int]],
    ) -> None:
        self.values = values
        self.ranges = ranges


def _reference_value_rows(profile: InsuranceMetadataProfileRevision) -> _ReferenceRows:
    rows: list[tuple[str, str, str, str | None]] = []
    ranges: list[tuple[str, int, int]] = []

    def append(
        kind: str,
        values: tuple[tuple[str, str, str | None], ...],
    ) -> None:
        start = len(rows) + 6
        rows.extend((kind, code, label, replacement) for code, label, replacement in values)
        ranges.append((kind, start, len(rows) + 5))

    append(
        "authority",
        tuple(
            (value.code, value.label, value.replacement_code)
            for value in profile.authority_values
        ),
    )
    append(
        "taxonomy_id",
        ((profile.taxonomy_id, profile.taxonomy_id, None),),
    )
    append(
        "taxonomy_revision_id",
        ((profile.taxonomy_revision_id, profile.taxonomy_revision_id, None),),
    )
    append(
        "precedence_policy_revision_id",
        (
            (
                profile.precedence_policy_revision_id,
                profile.precedence_policy_revision_id,
                None,
            ),
        ),
    )
    append(
        "precedence_authority_tier",
        tuple(
            (value.code, value.label, value.replacement_code)
            for value in profile.precedence_authority_tier_values
        ),
    )
    if any(end < start for _kind, start, end in ranges):
        raise ValueError("Workbook Profile requires labeled reference values")
    return _ReferenceRows(values=rows, ranges=ranges)


def _build_workbook(
    *,
    manifest: MetadataWorkbookExportManifestV2,
    profile: InsuranceMetadataProfileRevision,
    review_set: InsuranceMetadataReviewSet,
    rule_units: tuple[WorkbookRuleUnitInventoryItem, ...],
    reference_rows: _ReferenceRows,
) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    defaults = workbook.create_sheet("Document Defaults")
    overrides = workbook.create_sheet("Rule Unit Overrides")
    references = workbook.create_sheet("Reference Values")
    hidden_manifest = workbook.create_sheet("_Manifest")

    _write_instructions(instructions, manifest)
    _write_reference_values(references, reference_rows)
    _write_document_default(defaults, review_set.reviews[0])
    _write_rule_units(overrides, review_set, rule_units)
    _apply_validations(defaults, overrides, reference_rows, len(rule_units))
    _write_manifest(hidden_manifest, manifest)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
    hidden_manifest.sheet_state = "hidden"
    hidden_manifest.protection.sheet = True
    workbook.properties.title = "Insurance Rule Metadata Review"
    workbook.properties.subject = TEMPLATE_REVISION
    workbook.properties.creator = "Proof Agent"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


_TITLE_FILL = PatternFill("solid", fgColor="172554")
_HEADER_FILL = PatternFill("solid", fgColor="DBEAFE")
_EDITABLE_FILL = PatternFill("solid", fgColor="FEF3C7")
_READ_ONLY_FILL = PatternFill("solid", fgColor="F1F5F9")
_THIN_BOTTOM = Border(bottom=Side(style="thin", color="CBD5E1"))


def _write_instructions(sheet: Any, manifest: MetadataWorkbookExportManifestV2) -> None:
    sheet["A1"] = "Insurance Metadata Review Workbook"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = _TITLE_FILL
    sheet.merge_cells("A1:F1")
    rows = (
        ("Template", manifest.template_revision),
        ("Source", manifest.source_id),
        ("Document revision", f"{manifest.document_id} / {manifest.revision_id}"),
        ("Profile revision", manifest.profile_revision_id),
        ("Expires", manifest.expires_at.isoformat()),
        ("Workflow", "Edit yellow cells, save the file, then upload it for Preview and Apply."),
        ("Safety", "Do not add formulas, sheets, columns, external links, or embedded content."),
    )
    for row, (label, value) in enumerate(rows, 3):
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 88
    sheet.freeze_panes = "A3"


_DEFAULT_HEADERS = (
    "review_id",
    "review_identity",
    "review_version",
    "document_id",
    "revision_id",
    "structured_build_id",
    "profile_revision_id",
    "state",
    "scope",
    *GOVERNED_FIELDS,
)


def _write_document_default(sheet: Any, review: InsuranceMetadataReviewV2) -> None:
    _sheet_title(sheet, "Document Metadata Default", "One required default for every Rule Unit.")
    _write_headers(sheet, _DEFAULT_HEADERS)
    values = _metadata_values(review.current_draft)
    row = (
        review.review_id,
        review.review_identity,
        review.review_version,
        review.document_id,
        review.revision_id,
        review.structured_build_id,
        review.profile_revision_id,
        review.state,
        review.scope,
        *_value_tuple(values),
    )
    _write_row(sheet, 6, row, editable_from=10)
    sheet.freeze_panes = "J6"
    sheet.auto_filter.ref = f"A5:{get_column_letter(len(_DEFAULT_HEADERS))}6"
    sheet.protection.sheet = True
    sheet.protection.autoFilter = False


_OVERRIDE_HEADERS = (
    "canonical_anchor",
    "citation_uri",
    "safe_preview",
    "review_id",
    "review_identity",
    "review_version",
    "state",
    "current",
    "override_mode",
    "override_reason",
    *GOVERNED_FIELDS,
)


def _write_rule_units(
    sheet: Any,
    review_set: InsuranceMetadataReviewSet,
    rule_units: tuple[WorkbookRuleUnitInventoryItem, ...],
) -> None:
    _sheet_title(
        sheet,
        "Rule Unit Metadata Overrides",
        "All Rule Units are listed. Use override only for a business exception.",
    )
    _write_headers(sheet, _OVERRIDE_HEADERS)
    default = review_set.reviews[0]
    by_anchor = {
        review.canonical_anchor: review
        for review in review_set.reviews
        if review.scope == "rule_unit_override"
    }
    for row_number, item in enumerate(rule_units, 6):
        review = by_anchor.get(item.canonical_anchor)
        draft = default.current_draft if review is None else review.current_draft
        values = _metadata_values(draft)
        row = (
            item.canonical_anchor,
            item.citation_uri,
            item.safe_preview,
            None if review is None else review.review_id,
            None if review is None else review.review_identity,
            None if review is None else review.review_version,
            "inherited" if review is None else review.state,
            True if review is None else review.current,
            "inherit" if review is None else "override",
            None,
            *_value_tuple(values),
        )
        _write_row(sheet, row_number, row, editable_from=9)
    sheet.freeze_panes = "K6"
    if rule_units:
        sheet.auto_filter.ref = (
            f"A5:{get_column_letter(len(_OVERRIDE_HEADERS))}{len(rule_units) + 5}"
        )
    sheet.protection.sheet = True
    sheet.protection.autoFilter = False
    sheet.protection.sort = False


def _write_reference_values(sheet: Any, reference_rows: _ReferenceRows) -> None:
    headers = ("value_type", "code", "label", "replacement_code")
    _sheet_title(sheet, "Reference Values", "Published Profile values used by dropdowns.")
    _write_headers(sheet, headers)
    for row_number, values in enumerate(reference_rows.values, 6):
        for column, value in enumerate(values, 1):
            sheet.cell(row=row_number, column=column, value=value)
    sheet.freeze_panes = "A6"
    sheet.protection.sheet = True
    for column, width in enumerate((32, 34, 38, 34), 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _apply_validations(
    defaults: Any,
    overrides: Any,
    reference_rows: _ReferenceRows,
    rule_count: int,
) -> None:
    ranges = {kind: (start, end) for kind, start, end in reference_rows.ranges}

    def add(sheet: Any, column: str, kind: str, start_row: int, end_row: int) -> None:
        start, end = ranges[kind]
        validation = DataValidation(
            type="list",
            formula1=f"'Reference Values'!$B${start}:$B${end}",
            allow_blank=False,
        )
        validation.error = "Select a value from the published Metadata Profile."
        validation.errorTitle = "Profile value required"
        validation.prompt = "Choose a published Profile code."
        validation.promptTitle = "Governed metadata"
        sheet.add_data_validation(validation)
        validation.add(f"{column}{start_row}:{column}{end_row}")

    add(defaults, "J", "authority", 6, 6)
    add(defaults, "M", "taxonomy_id", 6, 6)
    add(defaults, "N", "taxonomy_revision_id", 6, 6)
    add(defaults, "O", "precedence_policy_revision_id", 6, 6)
    add(defaults, "P", "precedence_authority_tier", 6, 6)
    if rule_count:
        end = rule_count + 5
        add(overrides, "K", "authority", 6, end)
        add(overrides, "N", "taxonomy_id", 6, end)
        add(overrides, "O", "taxonomy_revision_id", 6, end)
        add(overrides, "P", "precedence_policy_revision_id", 6, end)
        add(overrides, "Q", "precedence_authority_tier", 6, end)
        mode = DataValidation(type="list", formula1='"inherit,override"', allow_blank=False)
        overrides.add_data_validation(mode)
        mode.add(f"I6:I{end}")


def _write_manifest(sheet: Any, manifest: MetadataWorkbookExportManifestV2) -> None:
    keys = (
        "schema_version",
        "template_revision",
        "export_id",
        "environment_id",
        "source_id",
        "document_id",
        "revision_id",
        "structured_build_id",
        "profile_revision_id",
        "review_set_id",
        "review_set_identity",
        "review_set_generation",
        "exported_at",
        "expires_at",
        "canonical_anchor_set_sha256",
    )
    payload = manifest.model_dump(mode="json")
    sheet.append(("manifest_key", "manifest_value"))
    for key in keys:
        sheet.append((key, payload[key]))
    sheet.append(())
    sheet.append(
        (
            "scope",
            "canonical_anchor",
            "override_mode",
            "review_id",
            "review_identity",
            "review_version",
            *GOVERNED_FIELDS,
        )
    )
    for base in manifest.base_reviews:
        sheet.append(
            (
                base.scope,
                base.canonical_anchor,
                base.override_mode,
                base.review_id,
                base.review_identity,
                base.review_version,
                *_value_tuple(base.values),
            )
        )
    sheet.append(())
    sheet.append(("registered_validation_range",))
    for registered in manifest.registered_validation_ranges:
        sheet.append((registered,))


def _sheet_title(sheet: Any, title: str, subtitle: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = _TITLE_FILL
    sheet.merge_cells("A1:H1")
    sheet["A3"] = subtitle
    sheet["A3"].font = Font(color="475569", italic=True)
    sheet.merge_cells("A3:H3")


def _write_headers(sheet: Any, headers: tuple[str, ...]) -> None:
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="172554")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = _THIN_BOTTOM
        sheet.column_dimensions[get_column_letter(column)].width = min(
            max(len(header) + 3, 14), 34
        )
    sheet.row_dimensions[5].height = 34


def _write_row(
    sheet: Any,
    row_number: int,
    values: tuple[object, ...],
    *,
    editable_from: int,
) -> None:
    for column, value in enumerate(values, 1):
        cell = sheet.cell(row=row_number, column=column, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=column in {2, 3})
        cell.fill = _EDITABLE_FILL if column >= editable_from else _READ_ONLY_FILL
        cell.protection = Protection(locked=column < editable_from)
        if isinstance(value, date) and not isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd"


def _base_review(
    review: InsuranceMetadataReviewV2,
    *,
    override_mode: Literal["default", "inherit", "override"],
) -> WorkbookReviewBaseV2:
    return WorkbookReviewBaseV2(
        scope=review.scope,
        canonical_anchor=review.canonical_anchor,
        override_mode=override_mode,
        review_id=review.review_id,
        review_identity=review.review_identity,
        review_version=review.review_version,
        values=_metadata_values(review.current_draft),
    )


def _metadata_values(draft: InsuranceRuleMetadataDraft) -> WorkbookMetadataValuesV2:
    applicability = draft.applicability
    precedence = draft.precedence
    return WorkbookMetadataValuesV2(
        authority=draft.authority,
        effective_from=draft.effective_from,
        effective_to=draft.effective_to,
        taxonomy_id=None if applicability is None else applicability.taxonomy_id,
        taxonomy_revision_id=(
            None if applicability is None else applicability.taxonomy_revision_id
        ),
        precedence_policy_revision_id=(
            None if precedence is None else precedence.policy_revision_id
        ),
        precedence_authority_tier=(
            None if precedence is None else precedence.authority_tier
        ),
        precedence_order=None if precedence is None else precedence.order,
    )


def _value_tuple(values: WorkbookMetadataValuesV2) -> tuple[object, ...]:
    return tuple(getattr(values, field) for field in GOVERNED_FIELDS)


def _sha256(value: object) -> str:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


__all__ = [
    "GOVERNED_FIELDS",
    "MetadataWorkbookExportManifestV2",
    "MetadataWorkbookExportAuthorityV2",
    "MetadataWorkbookExportV2",
    "MetadataWorkbookApplyCommitV2",
    "MetadataWorkbookApplyResultV2",
    "MetadataWorkbookImportPreviewV2",
    "MetadataWorkbookImportPreviewAuthorityV2",
    "MetadataWorkbookValidationError",
    "MetadataWorkbookValidationIssueV2",
    "MetadataWorkbookValidationReportV2",
    "SHEET_NAMES",
    "TEMPLATE_REVISION",
    "WorkbookMetadataValuesV2",
    "WorkbookFieldMergeV2",
    "WorkbookOverrideModeMergeV2",
    "WorkbookReviewBaseV2",
    "WorkbookRuleUnitInventoryItem",
    "apply_metadata_workbook_import_preview_v2",
    "create_metadata_workbook_import_preview_v2",
    "generate_metadata_workbook_v2",
]
