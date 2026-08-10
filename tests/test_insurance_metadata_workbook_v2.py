from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
import pytest

from proof_agent.capabilities.knowledge.hybrid.metadata_review import (
    advance_insurance_metadata_review_set,
    create_insurance_metadata_review_set,
    proofagent_insurance_reference_profile,
    save_metadata_review_draft,
)
from proof_agent.capabilities.knowledge.hybrid.metadata_workbook import (
    WorkbookRuleUnitInventoryItem,
    MetadataWorkbookValidationError,
    apply_metadata_workbook_import_preview_v2,
    create_metadata_workbook_import_preview_v2,
    generate_metadata_workbook_v2,
)
from proof_agent.capabilities.knowledge.hybrid.workbook import (
    InsuranceMetadataDraftInput,
)
from proof_agent.contracts.insurance_rules import (
    InsuranceRuleApplicability,
    InsuranceRuleMetadataDraft,
    InsuranceRulePrecedence,
)


def _complete_default() -> InsuranceRuleMetadataDraft:
    return InsuranceRuleMetadataDraft(
        metadata_draft_id="document-default-proposal",
        document_id="document-1",
        revision_id="revision-1",
        authority="national",
        effective_from=date(2026, 1, 1),
        applicability=InsuranceRuleApplicability(
            taxonomy_id="insurance-product-applicability",
            taxonomy_revision_id="taxonomy-2026-01",
        ),
        precedence=InsuranceRulePrecedence(
            policy_revision_id="precedence-2026-01",
            authority_tier="policy_terms",
            order=10,
        ),
    )


def test_generate_export_is_a_profile_bound_five_sheet_workbook() -> None:
    profile = proofagent_insurance_reference_profile()
    review_set = create_insurance_metadata_review_set(
        source_id="ks_insurance",
        structured_build_id="build-1",
        profile=profile,
        document_default=_complete_default(),
        parser_proposals=(
            InsuranceMetadataDraftInput(
                metadata_draft_id="parser-override-1",
                origin="pdf",
                source_id="ks_insurance",
                document_id="document-1",
                revision_id="revision-1",
                canonical_anchor="page:1#rule:coverage",
            ),
            InsuranceMetadataDraftInput(
                metadata_draft_id="parser-inherited-2",
                origin="pdf",
                source_id="ks_insurance",
                document_id="document-1",
                revision_id="revision-1",
                canonical_anchor="page:2#rule:exclusions",
                authority="national",
                effective_from=date(2026, 1, 1),
                taxonomy_id="insurance-product-applicability",
                taxonomy_revision_id="taxonomy-2026-01",
                precedence_policy_revision_id="precedence-2026-01",
                precedence_authority_tier="policy_terms",
                precedence_order=10,
            ),
        ),
        canonical_anchors=(
            "page:1#rule:coverage",
            "page:2#rule:exclusions",
        ),
    )
    exported_at = datetime(2026, 8, 8, tzinfo=UTC)

    exported = generate_metadata_workbook_v2(
        export_id="workbook-export-1",
        environment_id="disposable-local.v1",
        review_set=review_set,
        profile=profile,
        rule_units=(
            WorkbookRuleUnitInventoryItem(
                canonical_anchor="page:1#rule:coverage",
                citation_uri="proof://knowledge/ks_insurance/document-1/page:1",
                safe_preview="Coverage follows the signed policy terms.",
            ),
            WorkbookRuleUnitInventoryItem(
                canonical_anchor="page:2#rule:exclusions",
                citation_uri="proof://knowledge/ks_insurance/document-1/page:2",
                safe_preview="Exclusions are listed in the policy schedule.",
            ),
        ),
        exported_at=exported_at,
        expires_at=exported_at + timedelta(days=30),
    )

    assert exported.manifest.template_revision == "insurance-rule-metadata.v2"
    assert exported.manifest.review_set_identity == review_set.review_set_identity
    assert exported.manifest.profile_revision_id == profile.profile_revision_id
    assert exported.manifest.canonical_anchors == (
        "page:1#rule:coverage",
        "page:2#rule:exclusions",
    )
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    assert workbook.sheetnames == [
        "Instructions",
        "Document Defaults",
        "Rule Unit Overrides",
        "Reference Values",
        "_Manifest",
    ]
    assert workbook["_Manifest"].sheet_state == "hidden"
    assert workbook["Document Defaults"].protection.sheet is True
    assert workbook["Rule Unit Overrides"].protection.sheet is True

    default_sheet = workbook["Document Defaults"]
    assert default_sheet["A5"].value == "review_id"
    assert default_sheet["A6"].value == review_set.reviews[0].review_id
    assert default_sheet["J6"].value == "national"
    assert default_sheet["J6"].protection.locked is False

    override_sheet = workbook["Rule Unit Overrides"]
    assert override_sheet["A5"].value == "canonical_anchor"
    assert [override_sheet[f"A{row}"].value for row in (6, 7)] == [
        "page:1#rule:coverage",
        "page:2#rule:exclusions",
    ]
    assert [override_sheet[f"I{row}"].value for row in (6, 7)] == [
        "override",
        "inherit",
    ]
    assert override_sheet["K6"].protection.locked is False
    assert len(override_sheet.data_validations.dataValidation) >= 2
    validation_formulas = {
        validation.formula1
        for validation in override_sheet.data_validations.dataValidation
    }
    assert "'Reference Values'!$B$6:$B$8" in validation_formulas
    assert "'Reference Values'!$B$12:$B$15" in validation_formulas
    assert set(exported.manifest.registered_validation_ranges) >= {
        "'Reference Values'!$B$6:$B$8",
        "'Reference Values'!$B$12:$B$15",
    }

    formulas = [
        cell.coordinate
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert formulas == []


def test_import_preview_three_way_merges_non_overlapping_changes() -> None:
    profile = proofagent_insurance_reference_profile()
    review_set = create_insurance_metadata_review_set(
        source_id="ks_insurance",
        structured_build_id="build-1",
        profile=profile,
        document_default=_complete_default(),
        parser_proposals=(),
        canonical_anchors=(),
    )
    exported_at = datetime(2026, 8, 8, tzinfo=UTC)
    exported = generate_metadata_workbook_v2(
        export_id="workbook-export-merge",
        environment_id="disposable-local.v1",
        review_set=review_set,
        profile=profile,
        rule_units=(),
        exported_at=exported_at,
        expires_at=exported_at + timedelta(days=30),
    )
    returned = load_workbook(BytesIO(exported.content), data_only=False)
    returned["Document Defaults"]["J6"] = "provincial"
    returned_bytes = BytesIO()
    returned.save(returned_bytes)

    current_default = review_set.reviews[0]
    server_change = save_metadata_review_draft(
        current_default,
        profile=profile,
        expected_review_version=current_default.review_version,
        expected_review_identity=current_default.review_identity,
        actor="dashboard-operator",
        reason="Updated effective period from signed policy.",
        changes={"effective_from": date(2026, 2, 1)},
    )
    current_set = advance_insurance_metadata_review_set(
        review_set,
        server_change.review,
    )

    preview = create_metadata_workbook_import_preview_v2(
        preview_id="workbook-preview-1",
        export_manifest=exported.manifest,
        returned_content=returned_bytes.getvalue(),
        current_review_set=current_set,
        profile=profile,
        previewed_at=exported_at + timedelta(hours=1),
    )

    by_field = {
        merge.field: merge
        for merge in preview.field_merges
        if merge.scope == "document_default"
    }
    assert preview.state == "ready_to_apply"
    assert preview.conflict_count == 0
    assert by_field["authority"].classification == "workbook_only"
    assert by_field["authority"].proposed_value == "provincial"
    assert by_field["effective_from"].classification == "server_only"
    assert by_field["effective_from"].proposed_value == date(2026, 2, 1)
    assert current_set.reviews[0].current_draft.authority == "national"

    applied = apply_metadata_workbook_import_preview_v2(
        preview,
        current_review_set=current_set,
        profile=profile,
        expected_preview_identity=preview.preview_identity,
        actor="workbook-operator",
        reason="Apply reviewed bulk metadata changes.",
    )

    assert applied.review_set.generation == current_set.generation + 1
    assert applied.review_set.reviews[0].current_draft.authority == "provincial"
    assert applied.review_set.reviews[0].current_draft.effective_from == date(
        2026, 2, 1
    )
    assert [decision.action for decision in applied.decisions] == ["workbook_apply"]
    assert current_set.reviews[0].current_draft.authority == "national"


def test_import_preview_surfaces_same_field_conflict_and_blocks_apply() -> None:
    profile = proofagent_insurance_reference_profile()
    review_set = create_insurance_metadata_review_set(
        source_id="ks_insurance",
        structured_build_id="build-1",
        profile=profile,
        document_default=_complete_default(),
        parser_proposals=(),
        canonical_anchors=(),
    )
    exported_at = datetime(2026, 8, 8, tzinfo=UTC)
    exported = generate_metadata_workbook_v2(
        export_id="workbook-export-conflict",
        environment_id="disposable-local.v1",
        review_set=review_set,
        profile=profile,
        rule_units=(),
        exported_at=exported_at,
        expires_at=exported_at + timedelta(days=30),
    )
    returned = load_workbook(BytesIO(exported.content), data_only=False)
    returned["Document Defaults"]["J6"] = "provincial"
    returned_bytes = BytesIO()
    returned.save(returned_bytes)
    default = review_set.reviews[0]
    server_change = save_metadata_review_draft(
        default,
        profile=profile,
        expected_review_version=default.review_version,
        expected_review_identity=default.review_identity,
        actor="dashboard-operator",
        reason="Use institution authority.",
        changes={"authority": "institution"},
    )
    current_set = advance_insurance_metadata_review_set(
        review_set,
        server_change.review,
    )

    preview = create_metadata_workbook_import_preview_v2(
        preview_id="workbook-preview-conflict",
        export_manifest=exported.manifest,
        returned_content=returned_bytes.getvalue(),
        current_review_set=current_set,
        profile=profile,
        previewed_at=exported_at + timedelta(hours=1),
    )

    conflict = next(merge for merge in preview.field_merges if merge.field == "authority")
    assert preview.state == "conflicts"
    assert preview.conflict_count == 1
    assert conflict.classification == "conflict"
    assert conflict.base_value == "national"
    assert conflict.server_value == "institution"
    assert conflict.workbook_value == "provincial"
    with pytest.raises(Exception, match="unresolved conflicts"):
        apply_metadata_workbook_import_preview_v2(
            preview,
            current_review_set=current_set,
            profile=profile,
            expected_preview_identity=preview.preview_identity,
            actor="workbook-operator",
            reason="Must not partially apply.",
        )


def test_import_preview_rejects_every_cell_formula() -> None:
    profile = proofagent_insurance_reference_profile()
    review_set = create_insurance_metadata_review_set(
        source_id="ks_insurance",
        structured_build_id="build-1",
        profile=profile,
        document_default=_complete_default(),
        parser_proposals=(),
        canonical_anchors=(),
    )
    exported_at = datetime(2026, 8, 8, tzinfo=UTC)
    exported = generate_metadata_workbook_v2(
        export_id="workbook-export-formula",
        environment_id="disposable-local.v1",
        review_set=review_set,
        profile=profile,
        rule_units=(),
        exported_at=exported_at,
        expires_at=exported_at + timedelta(days=30),
    )
    returned = load_workbook(BytesIO(exported.content), data_only=False)
    returned["Document Defaults"]["J6"] = '=HYPERLINK("https://invalid.example")'
    returned_bytes = BytesIO()
    returned.save(returned_bytes)

    with pytest.raises(
        MetadataWorkbookValidationError,
        match="metadata_workbook_formula_forbidden",
    ):
        create_metadata_workbook_import_preview_v2(
            preview_id="workbook-preview-formula",
            export_manifest=exported.manifest,
            returned_content=returned_bytes.getvalue(),
            current_review_set=review_set,
            profile=profile,
            previewed_at=exported_at + timedelta(hours=1),
        )


def test_workbook_apply_can_create_an_explicit_rule_unit_override() -> None:
    profile = proofagent_insurance_reference_profile()
    review_set = create_insurance_metadata_review_set(
        source_id="ks_insurance",
        structured_build_id="build-1",
        profile=profile,
        document_default=_complete_default(),
        parser_proposals=(
            InsuranceMetadataDraftInput(
                metadata_draft_id="matching-proposal",
                origin="pdf",
                source_id="ks_insurance",
                document_id="document-1",
                revision_id="revision-1",
                canonical_anchor="page:1#rule:coverage",
                authority="national",
                effective_from=date(2026, 1, 1),
                taxonomy_id="insurance-product-applicability",
                taxonomy_revision_id="taxonomy-2026-01",
                precedence_policy_revision_id="precedence-2026-01",
                precedence_authority_tier="policy_terms",
                precedence_order=10,
            ),
        ),
        canonical_anchors=("page:1#rule:coverage",),
    )
    assert len(review_set.reviews) == 1
    exported_at = datetime(2026, 8, 8, tzinfo=UTC)
    exported = generate_metadata_workbook_v2(
        export_id="workbook-export-new-override",
        environment_id="disposable-local.v1",
        review_set=review_set,
        profile=profile,
        rule_units=(
            WorkbookRuleUnitInventoryItem(
                canonical_anchor="page:1#rule:coverage",
                citation_uri="proof://knowledge/ks_insurance/document-1/page:1",
                safe_preview="Coverage follows the signed policy terms.",
            ),
        ),
        exported_at=exported_at,
        expires_at=exported_at + timedelta(days=30),
    )
    returned = load_workbook(BytesIO(exported.content), data_only=False)
    returned["Rule Unit Overrides"]["I6"] = "override"
    returned["Rule Unit Overrides"]["J6"] = "Provincial exception applies."
    returned["Rule Unit Overrides"]["K6"] = "provincial"
    returned_bytes = BytesIO()
    returned.save(returned_bytes)
    preview = create_metadata_workbook_import_preview_v2(
        preview_id="workbook-preview-new-override",
        export_manifest=exported.manifest,
        returned_content=returned_bytes.getvalue(),
        current_review_set=review_set,
        profile=profile,
        previewed_at=exported_at + timedelta(hours=1),
    )

    applied = apply_metadata_workbook_import_preview_v2(
        preview,
        current_review_set=review_set,
        profile=profile,
        expected_preview_identity=preview.preview_identity,
        actor="workbook-operator",
        reason="Apply the reviewed provincial exception.",
    )

    assert len(applied.review_set.reviews) == 2
    override = applied.review_set.reviews[1]
    assert override.scope == "rule_unit_override"
    assert override.canonical_anchor == "page:1#rule:coverage"
    assert override.parser_proposal.authority == "national"
    assert override.current_draft.authority == "provincial"
    assert applied.decisions[0].action == "workbook_apply"
