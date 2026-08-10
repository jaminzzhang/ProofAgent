from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from proof_agent.capabilities.knowledge.hybrid.metadata_review import (
    InsuranceMetadataProfileRevision,
    InsuranceMetadataProfileValue,
    create_insurance_metadata_review_set,
)
from proof_agent.capabilities.knowledge.hybrid.metadata_workbook import (
    create_metadata_workbook_import_preview_v2,
    generate_metadata_workbook_v2,
)
from proof_agent.capabilities.knowledge.hybrid.metadata_workbook_jobs import (
    MetadataWorkbookJobV2,
)
from proof_agent.capabilities.knowledge.ingestion.metadata_workbook_worker import (
    MetadataWorkbookV2Worker,
)
from proof_agent.capabilities.persistence.postgres.bundle import PostgresPersistenceBundle
from proof_agent.capabilities.persistence.postgres.database import upgrade_database
from proof_agent.capabilities.persistence.postgres.prepared_knowledge_publication_repository import (
    PostgresPreparedKnowledgePublicationRepository,
)
from proof_agent.configuration.hybrid_knowledge_repository import (
    FileSystemKnowledgeArtifactStore,
)
from proof_agent.contracts import (
    KnowledgeSource,
    KnowledgeSourceLifecycleState,
    KnowledgeSourceOperation,
    PreparedHybridKnowledgePublication,
)
from proof_agent.contracts.insurance_rules import (
    InsuranceRuleApplicability,
    InsuranceRuleMetadataDraft,
    InsuranceRulePrecedence,
)
from proof_agent.contracts.knowledge_index import ExactArtifactRef


pytestmark = pytest.mark.postgres_integration
pytest_plugins = ("postgres_fixtures",)


def test_postgres_workbook_preview_apply_is_atomic_and_one_use(
    postgres_dsn: str,
) -> None:
    upgrade_database(postgres_dsn)
    bundle = PostgresPersistenceBundle.create(postgres_dsn)
    suffix = uuid4().hex
    source_id = f"ks_{suffix}"
    document_id = str(uuid4())
    revision_id = str(uuid4())
    now = datetime(2026, 8, 8, tzinfo=UTC)
    profile = InsuranceMetadataProfileRevision(
        profile_id=f"insurance-authority-{suffix}",
        profile_revision_id=f"insurance-authority-{suffix}.v1",
        authority_codes=("national", "provincial"),
        authority_values=(
            InsuranceMetadataProfileValue(code="national", label="National"),
            InsuranceMetadataProfileValue(code="provincial", label="Provincial"),
        ),
        taxonomy_id="insurance-product-applicability",
        taxonomy_revision_id="taxonomy-2026-01",
        precedence_policy_revision_id="precedence-2026-01",
        precedence_authority_tiers=("policy_terms",),
        precedence_authority_tier_values=(
            InsuranceMetadataProfileValue(code="policy_terms", label="Policy terms"),
        ),
    )
    try:
        source = KnowledgeSource(
            source_id=source_id,
            name="Insurance terms",
            provider="hybrid_index",
            lifecycle_state=KnowledgeSourceLifecycleState.ACTIVE,
            params={},
            source_draft_version_id=str(uuid4()),
            created_at=now.isoformat(),
            updated_at=now.isoformat(),
        )
        bundle.knowledge.save_source(source, expected_revision=0)
        bundle.metadata_reviews.publish_profile(
            profile,
            display_name="Insurance authority",
            actor="profile-publisher",
            published_at=now,
        )
        bundle.metadata_reviews.bind_source_profile(
            source_id=source_id,
            profile_revision_id=profile.profile_revision_id,
            actor="source-editor",
            bound_at=now,
            production=True,
        )
        default = InsuranceRuleMetadataDraft(
            metadata_draft_id="document-default-proposal",
            document_id=document_id,
            revision_id=revision_id,
            authority="national",
            effective_from=date(2026, 1, 1),
            applicability=InsuranceRuleApplicability(
                taxonomy_id=profile.taxonomy_id,
                taxonomy_revision_id=profile.taxonomy_revision_id,
            ),
            precedence=InsuranceRulePrecedence(
                policy_revision_id=profile.precedence_policy_revision_id,
                authority_tier="policy_terms",
                order=10,
            ),
        )
        review_set = create_insurance_metadata_review_set(
            source_id=source_id,
            structured_build_id="build-1",
            profile=profile,
            document_default=default,
            parser_proposals=(),
            canonical_anchors=(),
        )
        bundle.metadata_reviews.put_review_set(review_set)
        exported = generate_metadata_workbook_v2(
            export_id=f"workbook-export-{suffix}",
            environment_id="production-private-plane.v1",
            review_set=review_set,
            profile=profile,
            rule_units=(),
            exported_at=now,
            expires_at=now + timedelta(days=30),
        )
        export_ref = ExactArtifactRef(
            artifact_uri=f"s3://proof-agent-test/workbooks/{suffix}.xlsx",
            version_id=f"sha256:{'1' * 64}",
            sha256="1" * 64,
            size_bytes=len(exported.content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        stored_export = bundle.metadata_workbooks.put_export(
            exported.manifest,
            artifact_ref=export_ref,
            actor="source-editor",
        )
        assert stored_export.state == "available"

        returned = load_workbook(BytesIO(exported.content), data_only=False)
        returned["Document Defaults"]["J6"] = "provincial"
        returned_bytes = BytesIO()
        returned.save(returned_bytes)
        preview = create_metadata_workbook_import_preview_v2(
            preview_id=f"workbook-preview-{suffix}",
            export_manifest=exported.manifest,
            returned_content=returned_bytes.getvalue(),
            current_review_set=review_set,
            profile=profile,
            previewed_at=now + timedelta(hours=1),
        )
        returned_ref = ExactArtifactRef(
            artifact_uri=f"s3://proof-agent-test/workbook-returns/{suffix}.xlsx",
            version_id=f"sha256:{'2' * 64}",
            sha256="2" * 64,
            size_bytes=len(returned_bytes.getvalue()),
            media_type=export_ref.media_type,
        )
        stored_preview = bundle.metadata_workbooks.put_preview(
            preview,
            original_ref=returned_ref,
            actor="source-editor",
            expires_at=now + timedelta(days=30),
        )
        assert stored_preview.state == "ready_to_apply"
        prepared_publications = PostgresPreparedKnowledgePublicationRepository(
            bundle.engine
        )
        bundle.knowledge_source_operations.save(
            KnowledgeSourceOperation(
                operation_id=f"prepare-operation-{suffix}",
                source_id=source_id,
                command="prepare_publication",
                status="succeeded",
                stage="prepared",
                source_revision=1,
                poll_after_ms=1_000,
                outcome_code="publication_prepared",
                outcome_detail="Publication validation completed.",
                created_at=now.isoformat(),
                updated_at=now.isoformat(),
                completed_at=now.isoformat(),
            )
        )
        prepared_publications.save_prepared(
            PreparedHybridKnowledgePublication(
                validation_id=f"validation-{suffix}",
                operation_id=f"prepare-operation-{suffix}",
                attempt_id=f"publication-attempt-{suffix}",
                fencing_token=1,
                source_id=source_id,
                source_draft_version_id=source.source_draft_version_id,
                candidate_digest="a" * 64,
                generation_id=f"generation-{suffix}",
                manifest_sha256="b" * 64,
                staged_projection_id=f"projection-{suffix}",
                attestation_sha256="c" * 64,
                smoke_result_sha256="d" * 64,
                state="prepared",
                prepared_at=now.isoformat(),
            )
        )

        committed = bundle.metadata_workbooks.apply_preview(
            source_id=source_id,
            preview_id=preview.preview_id,
            expected_preview_identity=preview.preview_identity,
            actor="source-editor",
            reason="Apply reviewed Workbook changes.",
            applied_at=now + timedelta(hours=2),
        )

        assert committed.source_revision == 2
        assert committed.review_set.reviews[0].current_draft.authority == "provincial"
        assert committed.preview.state == "applied"
        assert prepared_publications.get(f"validation-{suffix}").state == "invalidated"
        assert bundle.metadata_workbooks.get_export(
            source_id=source_id,
            export_id=exported.manifest.export_id,
        ).state == "consumed"
        with pytest.raises(Exception, match="Preview"):
            bundle.metadata_workbooks.apply_preview(
                source_id=source_id,
                preview_id=preview.preview_id,
                expected_preview_identity=preview.preview_identity,
                actor="source-editor",
                reason="Must not apply twice.",
                applied_at=now + timedelta(hours=3),
            )
    finally:
        bundle.close()


def test_postgres_workbook_v2_jobs_are_leased_and_fenced(postgres_dsn: str) -> None:
    upgrade_database(postgres_dsn)
    bundle = PostgresPersistenceBundle.create(postgres_dsn)
    now = datetime(2026, 8, 8, tzinfo=UTC)
    source_id = f"ks_{uuid4().hex}"
    operation_id = f"operation-{uuid4().hex}"
    job_id = str(uuid4())
    try:
        bundle.knowledge.save_source(
            KnowledgeSource(
                source_id=source_id,
                name="Insurance terms",
                provider="hybrid_index",
                lifecycle_state=KnowledgeSourceLifecycleState.ACTIVE,
                params={},
                source_draft_version_id=str(uuid4()),
                created_at=now.isoformat(),
                updated_at=now.isoformat(),
            ),
            expected_revision=0,
        )
        bundle.knowledge_source_operations.save(
            KnowledgeSourceOperation(
                operation_id=operation_id,
                source_id=source_id,
                command="generate_metadata_workbook_export",
                status="queued",
                stage="metadata_workbook_export_queued",
                source_revision=1,
                poll_after_ms=500,
                created_at=now.isoformat(),
                updated_at=now.isoformat(),
            )
        )
        job = MetadataWorkbookJobV2(
            job_id=job_id,
            operation_id=operation_id,
            source_id=source_id,
            document_id=str(uuid4()),
            revision_id=str(uuid4()),
            source_revision=1,
            command="generate_export",
            resource_id=operation_id,
            request_sha256="d" * 64,
            state="READY",
            fencing_token=0,
            created_by="operator-1",
            created_at=now,
            updated_at=now,
        )

        assert bundle.metadata_workbooks.enqueue_job(job) == job
        claim = bundle.metadata_workbooks.claim_next_job(
            worker_id="knowledge-worker-1",
            lease_seconds=60,
        )
        assert claim is not None
        assert claim.fencing_token == 1
        assert bundle.metadata_workbooks.require_job_claim(claim).state == "CLAIMED"
        completed = bundle.metadata_workbooks.complete_job(
            claim,
            completed_at=claim.claimed_at + timedelta(seconds=1),
        )
        assert completed.state == "COMPLETED"
        assert bundle.metadata_workbooks.claim_next_job(
            worker_id="knowledge-worker-1",
            lease_seconds=60,
        ) is None
    finally:
        bundle.close()


def test_configuration_uow_exposes_workbook_authority_on_the_same_connection(
    postgres_dsn: str,
) -> None:
    upgrade_database(postgres_dsn)
    bundle = PostgresPersistenceBundle.create(postgres_dsn)
    try:
        with bundle.configuration_uow() as uow:
            assert uow.metadata_workbooks is not None
            assert uow.metadata_workbooks._connection_source is uow._connection
    finally:
        bundle.close()


def test_workbook_worker_generates_an_exact_export_and_completes_its_operation(
    postgres_dsn: str,
    tmp_path,
) -> None:
    upgrade_database(postgres_dsn)
    bundle = PostgresPersistenceBundle.create(postgres_dsn)
    artifacts = FileSystemKnowledgeArtifactStore(tmp_path / "workbook-artifacts")
    now = datetime.now(UTC)
    suffix = uuid4().hex
    source_id = f"ks_{suffix}"
    document_id = str(uuid4())
    revision_id = str(uuid4())
    operation_id = f"operation-{suffix}"
    profile = InsuranceMetadataProfileRevision(
        profile_id=f"insurance-authority-{suffix}",
        profile_revision_id=f"insurance-authority-{suffix}.v1",
        authority_codes=("national", "provincial"),
        authority_values=(
            InsuranceMetadataProfileValue(code="national", label="National"),
            InsuranceMetadataProfileValue(code="provincial", label="Provincial"),
        ),
        taxonomy_id="insurance-product-applicability",
        taxonomy_revision_id="taxonomy-2026-01",
        precedence_policy_revision_id="precedence-2026-01",
        precedence_authority_tiers=("policy_terms",),
        precedence_authority_tier_values=(
            InsuranceMetadataProfileValue(code="policy_terms", label="Policy terms"),
        ),
    )

    class EmptyInventory:
        def rule_units(self, **_kwargs):
            return ()

    try:
        bundle.knowledge.save_source(
            KnowledgeSource(
                source_id=source_id,
                name="Insurance terms",
                provider="hybrid_index",
                lifecycle_state=KnowledgeSourceLifecycleState.ACTIVE,
                params={},
                source_draft_version_id=str(uuid4()),
                created_at=now.isoformat(),
                updated_at=now.isoformat(),
            ),
            expected_revision=0,
        )
        bundle.metadata_reviews.publish_profile(
            profile,
            display_name="Insurance authority",
            actor="profile-publisher",
            published_at=now,
        )
        bundle.metadata_reviews.bind_source_profile(
            source_id=source_id,
            profile_revision_id=profile.profile_revision_id,
            actor="source-editor",
            bound_at=now,
            production=True,
        )
        review_set = create_insurance_metadata_review_set(
            source_id=source_id,
            structured_build_id="build-1",
            profile=profile,
            document_default=InsuranceRuleMetadataDraft(
                metadata_draft_id="document-default-proposal",
                document_id=document_id,
                revision_id=revision_id,
                authority="national",
                effective_from=date(2026, 1, 1),
                applicability=InsuranceRuleApplicability(
                    taxonomy_id=profile.taxonomy_id,
                    taxonomy_revision_id=profile.taxonomy_revision_id,
                ),
                precedence=InsuranceRulePrecedence(
                    policy_revision_id=profile.precedence_policy_revision_id,
                    authority_tier="policy_terms",
                    order=10,
                ),
            ),
            parser_proposals=(),
            canonical_anchors=(),
        )
        bundle.metadata_reviews.put_review_set(review_set)
        bundle.knowledge_source_operations.save(
            KnowledgeSourceOperation(
                operation_id=operation_id,
                source_id=source_id,
                command="generate_metadata_workbook_export",
                status="queued",
                stage="metadata_workbook_export_queued",
                source_revision=1,
                poll_after_ms=500,
                created_at=now.isoformat(),
                updated_at=now.isoformat(),
            )
        )
        bundle.metadata_workbooks.enqueue_job(
            MetadataWorkbookJobV2(
                job_id=str(uuid4()),
                operation_id=operation_id,
                source_id=source_id,
                document_id=document_id,
                revision_id=revision_id,
                source_revision=1,
                command="generate_export",
                resource_id=operation_id,
                request_sha256="e" * 64,
                state="READY",
                fencing_token=0,
                created_by="operator-1",
                created_at=now,
                updated_at=now,
            )
        )

        outcome = MetadataWorkbookV2Worker(
            jobs=bundle.metadata_workbooks,
            reviews=bundle.metadata_reviews,
            workbooks=bundle.metadata_workbooks,
            inventory=EmptyInventory(),
            unit_of_work_factory=bundle.configuration_uow,
            artifact_store=artifacts,
            environment_id="production-private-plane.v1",
            worker_id="knowledge-worker-1",
            clock=lambda: now + timedelta(seconds=1),
        ).run_once()

        assert outcome is not None
        assert outcome.state == "completed"
        export = bundle.metadata_workbooks.get_export(
            source_id=source_id,
            export_id=operation_id,
        )
        assert export is not None
        assert export.state == "available"
        content = artifacts.get_exact(export.artifact_ref)
        assert load_workbook(BytesIO(content), data_only=False).sheetnames == [
            "Instructions",
            "Document Defaults",
            "Rule Unit Overrides",
            "Reference Values",
            "_Manifest",
        ]
        operation = bundle.knowledge_source_operations.get(operation_id)
        assert operation is not None
        assert operation.status == "succeeded"
        assert operation.outcome_code == "metadata_workbook_export_completed"

        invalid_workbook = load_workbook(BytesIO(content), data_only=False)
        invalid_workbook["Document Defaults"]["J6"] = "=1+1"
        invalid_bytes = BytesIO()
        invalid_workbook.save(invalid_bytes)
        invalid_ref = artifacts.put_immutable(
            key=f"metadata-workbooks/v2/{operation_id}/invalid-return.xlsx",
            content=invalid_bytes.getvalue(),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        preview_operation_id = f"preview-operation-{suffix}"
        bundle.knowledge_source_operations.save(
            KnowledgeSourceOperation(
                operation_id=preview_operation_id,
                source_id=source_id,
                command="create_metadata_workbook_import_preview",
                status="queued",
                stage="metadata_workbook_preview_queued",
                source_revision=1,
                poll_after_ms=500,
                created_at=now.isoformat(),
                updated_at=now.isoformat(),
            )
        )
        bundle.metadata_workbooks.enqueue_job(
            MetadataWorkbookJobV2(
                job_id=str(uuid4()),
                operation_id=preview_operation_id,
                source_id=source_id,
                document_id=document_id,
                revision_id=revision_id,
                source_revision=1,
                command="create_preview",
                resource_id=preview_operation_id,
                parent_resource_id=operation_id,
                request_sha256="f" * 64,
                original_ref=invalid_ref,
                state="READY",
                fencing_token=0,
                created_by="operator-1",
                created_at=now,
                updated_at=now,
            )
        )

        invalid_outcome = MetadataWorkbookV2Worker(
            jobs=bundle.metadata_workbooks,
            reviews=bundle.metadata_reviews,
            workbooks=bundle.metadata_workbooks,
            inventory=EmptyInventory(),
            unit_of_work_factory=bundle.configuration_uow,
            artifact_store=artifacts,
            environment_id="production-private-plane.v1",
            worker_id="knowledge-worker-1",
            clock=lambda: now + timedelta(seconds=2),
        ).run_once()

        assert invalid_outcome is not None
        validation_operation = bundle.knowledge_source_operations.get(
            preview_operation_id
        )
        assert validation_operation is not None
        assert validation_operation.status == "failed"
        assert validation_operation.outcome_code == (
            "metadata_workbook_preview_validation_failed"
        )
        invalid_preview = bundle.metadata_workbooks.get_preview(
            source_id=source_id,
            preview_id=preview_operation_id,
        )
        assert invalid_preview is not None
        assert invalid_preview.state == "validation_failed"
        assert invalid_preview.validation_report.errors[0].code.startswith(
            "metadata_workbook_"
        )
    finally:
        artifacts.close()
        bundle.close()
