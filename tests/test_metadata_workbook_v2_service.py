from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from proof_agent.capabilities.knowledge.hybrid.metadata_review import (
    InsuranceMetadataProfileRevision,
    InsuranceMetadataProfileValue,
    create_insurance_metadata_review_set,
)
from proof_agent.capabilities.knowledge.ingestion.metadata_workbook_worker import (
    MetadataWorkbookV2Worker,
)
from proof_agent.capabilities.persistence.postgres.bundle import PostgresPersistenceBundle
from proof_agent.capabilities.persistence.postgres.database import upgrade_database
from proof_agent.control.knowledge.metadata_workbook_service import (
    KnowledgeSourceMetadataWorkbookService,
)
from proof_agent.control.knowledge.application import KnowledgeSourceCommandContext
from proof_agent.configuration.hybrid_knowledge_repository import (
    FileSystemKnowledgeArtifactStore,
)
from proof_agent.contracts import (
    AuditActorFacts,
    KnowledgeSource,
    KnowledgeSourceIntakeCapability,
    KnowledgeSourceLifecycleState,
    KnowledgeSourceProviderCapability,
    KnowledgeSourceProviderReadiness,
    Permission,
)
from proof_agent.contracts.insurance_rules import (
    InsuranceRuleApplicability,
    InsuranceRuleMetadataDraft,
    InsuranceRulePrecedence,
)


pytestmark = pytest.mark.postgres_integration
pytest_plugins = ("postgres_fixtures",)


class _Summary:
    def summary_for_source(self, _source_id: str) -> dict[str, int]:
        return {"ready": 1, "review_required": 1}


class _EmptyInventory:
    def rule_units(self, **_kwargs):
        return ()


def _provider() -> KnowledgeSourceProviderCapability:
    return KnowledgeSourceProviderCapability(
        provider="hybrid_index",
        creation_supported=True,
        intake=KnowledgeSourceIntakeCapability(
            content_types=("application/pdf",),
            max_file_bytes=50 * 1024 * 1024,
            max_batch_files=1,
            max_source_documents=10_000,
        ),
        features=("metadata_workbook_v2",),
        readiness=KnowledgeSourceProviderReadiness(state="ready"),
    )


def _actor() -> AuditActorFacts:
    return AuditActorFacts(
        subject="source-editor-1",
        identity_provider="enterprise-oidc",
        session_id=str(uuid4()),
        permissions=("knowledge_source.edit",),
    )


def _arrange_review_set(bundle: PostgresPersistenceBundle):
    now = datetime.now(UTC)
    suffix = uuid4().hex
    source_id = f"ks_{suffix}"
    document_id = str(uuid4())
    revision_id = str(uuid4())
    profile = InsuranceMetadataProfileRevision(
        profile_id=f"insurance-authority-{suffix}",
        profile_revision_id=f"insurance-authority-{suffix}.v1",
        authority_codes=("national", "provincial"),
        authority_values=(
            InsuranceMetadataProfileValue(code="national", label="National"),
            InsuranceMetadataProfileValue(code="provincial", label="Provincial"),
        ),
        taxonomy_id="insurance-product-applicability",
        taxonomy_revision_id="taxonomy-2026-01",
        precedence_policy_revision_id="precedence-2026-01",
        precedence_authority_tiers=("policy_terms",),
        precedence_authority_tier_values=(
            InsuranceMetadataProfileValue(code="policy_terms", label="Policy terms"),
        ),
    )
    bundle.knowledge.save_source(
        KnowledgeSource(
            source_id=source_id,
            name="Insurance terms",
            provider="hybrid_index",
            lifecycle_state=KnowledgeSourceLifecycleState.ACTIVE,
            params={},
            source_draft_version_id=str(uuid4()),
            created_at=now.isoformat(),
            updated_at=now.isoformat(),
        ),
        expected_revision=0,
    )
    bundle.metadata_reviews.publish_profile(
        profile,
        display_name="Insurance authority",
        actor="profile-publisher",
        published_at=now,
    )
    bundle.metadata_reviews.bind_source_profile(
        source_id=source_id,
        profile_revision_id=profile.profile_revision_id,
        actor="source-editor",
        bound_at=now,
        production=True,
    )
    review_set = create_insurance_metadata_review_set(
        source_id=source_id,
        structured_build_id="build-1",
        profile=profile,
        document_default=InsuranceRuleMetadataDraft(
            metadata_draft_id="document-default-proposal",
            document_id=document_id,
            revision_id=revision_id,
            authority="national",
            effective_from=date(2026, 1, 1),
            applicability=InsuranceRuleApplicability(
                taxonomy_id=profile.taxonomy_id,
                taxonomy_revision_id=profile.taxonomy_revision_id,
            ),
            precedence=InsuranceRulePrecedence(
                policy_revision_id=profile.precedence_policy_revision_id,
                authority_tier="policy_terms",
                order=10,
            ),
        ),
        parser_proposals=(),
        canonical_anchors=(),
    )
    bundle.metadata_reviews.put_review_set(review_set)
    return now, source_id, document_id, revision_id


def test_generate_export_admission_is_atomic_and_exactly_replayable(
    postgres_dsn: str,
) -> None:
    upgrade_database(postgres_dsn)
    bundle = PostgresPersistenceBundle.create(postgres_dsn)
    now, source_id, document_id, revision_id = _arrange_review_set(bundle)
    service = KnowledgeSourceMetadataWorkbookService(
        unit_of_work_factory=bundle.configuration_uow,
        provider_capability=_provider(),
        summary_reader=_Summary(),
        knowledge=bundle.knowledge,
        metadata_reviews=bundle.metadata_reviews,
        workbooks=bundle.metadata_workbooks,
        artifact_store=None,
        clock=lambda: now,
    )
    try:
        operation = service.generate_export(
            source_id=source_id,
            document_id=document_id,
            revision_id=revision_id,
            expected_revision=1,
            idempotency_key="export-1",
            actor=_actor(),
        )
        replay = service.generate_export(
            source_id=source_id,
            document_id=document_id,
            revision_id=revision_id,
            expected_revision=1,
            idempotency_key="export-1",
            actor=_actor(),
        )

        assert replay == operation
        assert operation.command == "generate_metadata_workbook_export"
        assert operation.stage == "metadata_workbook_export_queued"
        assert operation.source_revision == 1
        assert bundle.knowledge.get_source_record(source_id).revision == 1
        with bundle.engine.connect() as connection:
            job = connection.exec_driver_sql(
                "SELECT command, resource_id, source_revision, state "
                "FROM hybrid_metadata_workbook_jobs WHERE operation_id = %s",
                (operation.operation_id,),
            ).one()
        assert job == (
            "generate_export",
            operation.operation_id,
            operation.source_revision,
            "READY",
        )
    finally:
        bundle.close()


def test_generated_export_can_be_returned_as_an_async_three_way_preview(
    postgres_dsn: str,
    tmp_path,
) -> None:
    upgrade_database(postgres_dsn)
    bundle = PostgresPersistenceBundle.create(postgres_dsn)
    artifacts = FileSystemKnowledgeArtifactStore(tmp_path / "workbook-service-artifacts")
    now, source_id, document_id, revision_id = _arrange_review_set(bundle)
    service = KnowledgeSourceMetadataWorkbookService(
        unit_of_work_factory=bundle.configuration_uow,
        provider_capability=_provider(),
        summary_reader=_Summary(),
        knowledge=bundle.knowledge,
        metadata_reviews=bundle.metadata_reviews,
        workbooks=bundle.metadata_workbooks,
        artifact_store=artifacts,
        clock=lambda: now,
    )
    try:
        operation = service.generate_export(
            source_id=source_id,
            document_id=document_id,
            revision_id=revision_id,
            expected_revision=1,
            idempotency_key="export-download-1",
            actor=_actor(),
        )
        outcome = MetadataWorkbookV2Worker(
            jobs=bundle.metadata_workbooks,
            reviews=bundle.metadata_reviews,
            workbooks=bundle.metadata_workbooks,
            inventory=_EmptyInventory(),
            unit_of_work_factory=bundle.configuration_uow,
            artifact_store=artifacts,
            environment_id="production-private-plane.v1",
            worker_id="knowledge-worker-1",
            clock=lambda: now + timedelta(seconds=1),
        ).run_once()
        assert outcome is not None and outcome.state == "completed"

        content, filename = service.download_export(
            source_id=source_id,
            export_id=operation.operation_id,
            context=KnowledgeSourceCommandContext(
                operator_subject="source-editor-1",
                permissions=(Permission.KNOWLEDGE_SOURCE_EDIT,),
            ),
        )

        assert filename == f"{source_id}-{document_id}-metadata-v2.xlsx"
        workbook = load_workbook(BytesIO(content), data_only=False)
        assert workbook.sheetnames[-1] == "_Manifest"
        workbook["Document Defaults"]["J6"] = "provincial"
        returned = BytesIO()
        workbook.save(returned)

        preview_operation = service.create_import_preview(
            source_id=source_id,
            export_id=operation.operation_id,
            filename=filename,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            content=BytesIO(returned.getvalue()),
            expected_revision=1,
            idempotency_key="preview-1",
            actor=_actor(),
        )

        assert preview_operation.source_revision == 1
        assert bundle.knowledge.get_source_record(source_id).revision == 1
        preview_outcome = MetadataWorkbookV2Worker(
            jobs=bundle.metadata_workbooks,
            reviews=bundle.metadata_reviews,
            workbooks=bundle.metadata_workbooks,
            inventory=_EmptyInventory(),
            unit_of_work_factory=bundle.configuration_uow,
            artifact_store=artifacts,
            environment_id="production-private-plane.v1",
            worker_id="knowledge-worker-1",
            clock=lambda: now + timedelta(seconds=2),
        ).run_once()
        assert preview_outcome is not None and preview_outcome.state == "completed"

        preview = service.get_import_preview(
            source_id=source_id,
            preview_id=preview_operation.operation_id,
            context=KnowledgeSourceCommandContext(
                operator_subject="source-editor-1",
                permissions=(Permission.KNOWLEDGE_SOURCE_VIEW,),
            ),
        )
        assert preview.state == "ready_to_apply"
        assert preview.conflict_count == 0
        authority_merge = next(
            merge for merge in preview.field_merges if merge.field == "authority"
        )
        assert authority_merge.classification == "workbook_only"
        assert authority_merge.proposed_value == "provincial"

        apply_operation = service.apply_import_preview(
            source_id=source_id,
            preview_id=preview.preview_id,
            expected_preview_identity=preview.preview_identity,
            expected_revision=1,
            reason="Apply the reviewed bulk metadata changes.",
            idempotency_key="apply-1",
            actor=_actor(),
        )
        assert apply_operation.source_revision == 1
        assert bundle.knowledge.get_source_record(source_id).revision == 1

        apply_outcome = MetadataWorkbookV2Worker(
            jobs=bundle.metadata_workbooks,
            reviews=bundle.metadata_reviews,
            workbooks=bundle.metadata_workbooks,
            inventory=_EmptyInventory(),
            unit_of_work_factory=bundle.configuration_uow,
            artifact_store=artifacts,
            environment_id="production-private-plane.v1",
            worker_id="knowledge-worker-1",
            clock=lambda: now + timedelta(seconds=3),
        ).run_once()
        assert apply_outcome is not None and apply_outcome.state == "completed"
        assert bundle.knowledge.get_source_record(source_id).revision == 2
        completed_operation = bundle.knowledge_source_operations.get(
            apply_operation.operation_id
        )
        assert completed_operation is not None
        assert completed_operation.status == "succeeded"
        assert completed_operation.source_revision == 2
        applied = service.get_import_preview(
            source_id=source_id,
            preview_id=preview.preview_id,
            context=KnowledgeSourceCommandContext(
                operator_subject="source-editor-1",
                permissions=(Permission.KNOWLEDGE_SOURCE_VIEW,),
            ),
        )
        assert applied.state == "applied"
    finally:
        artifacts.close()
        bundle.close()
