"""S3-first deterministic CSV intake into typed immutable Dataset Revisions."""

from __future__ import annotations

from collections.abc import Mapping
import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO, StringIO
import json
import re
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
import pyarrow as pa  # type: ignore[import-untyped]
import pyarrow.parquet as pq  # type: ignore[import-untyped]

from knowledge_source_service.domain.artifacts import ExactArtifactReference
from knowledge_source_service.domain.identities import content_identifier, sha256_json
from knowledge_source_service.domain.knowledge_catalog import (
    DatasetSourceVersion,
    StructuredField,
    StructuredRecord,
    StructuredScalar,
    StructuredValueType,
)
from knowledge_source_service.domain.publications import PublishedDatasetSourceVersion
from knowledge_source_service.ports.artifacts import ImmutableArtifactStore
from knowledge_source_service.ports.knowledge_catalog import KnowledgeCatalogWriter


_AUTHORITY_ID = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,127}$")
_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_MAX_XLSX_PARTS = 2_000
_MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
_PARQUET_MEDIA_TYPE = "application/vnd.apache.parquet"
_MAX_PARQUET_ROW_GROUPS = 10_000


@dataclass(frozen=True)
class CsvDatasetIntakeCommand:
    knowledge_space_id: str
    knowledge_source_id: str
    display_filename: str
    content: bytes
    field_types: Mapping[str, StructuredValueType]


class _BoundedDatasetIntakeApplication:
    def __init__(
        self,
        *,
        artifacts: ImmutableArtifactStore,
        catalog: KnowledgeCatalogWriter,
        pipeline_revision: str,
        max_content_bytes: int,
        max_records: int,
    ) -> None:
        if not pipeline_revision.strip():
            raise ValueError("pipeline_revision must not be blank")
        if max_content_bytes < 1 or max_records < 1:
            raise ValueError("dataset intake bounds must be positive")
        self._artifacts = artifacts
        self._catalog = catalog
        self._pipeline_revision = pipeline_revision
        self._max_content_bytes = max_content_bytes
        self._max_records = max_records


class CsvDatasetIntakeApplication(_BoundedDatasetIntakeApplication):
    """Finalize a typed Dataset Revision before exposing its Source Version."""

    def create_source_version(
        self,
        command: CsvDatasetIntakeCommand,
    ) -> PublishedDatasetSourceVersion:
        _validate_authority_id(command.knowledge_space_id, "knowledge_space_id")
        _validate_authority_id(command.knowledge_source_id, "knowledge_source_id")
        if type(command.content) is not bytes or not command.content:
            raise ValueError("CSV content must be nonempty exact bytes")
        if len(command.content) > self._max_content_bytes:
            raise ValueError("CSV content exceeds the admitted size limit")
        if b"\x00" in command.content:
            raise ValueError("CSV content contains a forbidden NUL byte")
        try:
            text = command.content.decode("utf-8")
        except UnicodeDecodeError as error:
            raise ValueError("CSV content must be valid UTF-8") from error
        field_types = tuple(command.field_types.items())
        if not field_types or any(not field.strip() for field, _ in field_types):
            raise ValueError("CSV intake requires nonblank declared fields")
        if len({field for field, _ in field_types}) != len(field_types):
            raise ValueError("CSV field declarations must be unique")
        field_order = tuple(field for field, _ in field_types)
        reader = csv.DictReader(StringIO(text, newline=""))
        if tuple(reader.fieldnames or ()) != field_order:
            raise ValueError("CSV headers must exactly match the declared field order")
        rows = list(reader)
        if not rows:
            raise ValueError("CSV dataset must contain at least one record")
        if len(rows) > self._max_records:
            raise ValueError("CSV dataset exceeds the admitted record limit")

        original_digest = f"sha256:{sha256(command.content).hexdigest()}"
        schema_revision_id = content_identifier(
            "dataset-schema",
            sha256_json({"fields": field_types}),
        )
        processing_lineage_digest = sha256_json(
            {
                "schema": "dataset-processing-lineage.v1",
                "pipeline_revision": self._pipeline_revision,
                "format": "csv",
                "schema_revision_id": schema_revision_id,
            }
        )
        version_digest = sha256_json(
            {
                "knowledge_space_id": command.knowledge_space_id,
                "knowledge_source_id": command.knowledge_source_id,
                "original_digest": original_digest,
                "schema_revision_id": schema_revision_id,
                "processing_lineage_digest": processing_lineage_digest,
            }
        )
        source_version_id = content_identifier("source-version", version_digest)
        dataset_revision_id = content_identifier("dataset-revision", version_digest)
        records = tuple(
            _record_from_row(
                row_number=row_number,
                row=row,
                field_types=field_types,
                dataset_revision_id=dataset_revision_id,
            )
            for row_number, row in enumerate(rows, start=1)
        )
        version = DatasetSourceVersion(
            knowledge_space_id=command.knowledge_space_id,
            knowledge_source_id=command.knowledge_source_id,
            knowledge_source_version_id=source_version_id,
            dataset_revision_id=dataset_revision_id,
            schema_revision_id=schema_revision_id,
            field_order=field_order,
            records=records,
        )
        key_root = (
            f"spaces/{command.knowledge_space_id}/sources/{command.knowledge_source_id}/"
            f"versions/{source_version_id}"
        )
        original = self._put_and_verify(
            object_key=f"{key_root}/originals/{original_digest.removeprefix('sha256:')}.csv",
            content=command.content,
            media_type="text/csv",
        )
        canonical_content = _canonical_json_bytes(
            {
                "schema_version": "structured-dataset-revision.v1",
                "knowledge_source_version_id": source_version_id,
                "dataset_revision_id": dataset_revision_id,
                "schema_revision_id": schema_revision_id,
                "processing_lineage_digest": processing_lineage_digest,
                "fields": [
                    {"field": field, "value_type": value_type}
                    for field, value_type in field_types
                ],
                "records": [
                    {
                        "record_id": record.record_id,
                        "content_hash": record.content_hash,
                        "fields": [
                            {
                                "field": field.field,
                                "value_type": field.value_type,
                                "value": field.value,
                            }
                            for field in record.fields
                        ],
                    }
                    for record in records
                ],
            }
        )
        canonical_digest = f"sha256:{sha256(canonical_content).hexdigest()}"
        canonical = self._put_and_verify(
            object_key=(
                f"{key_root}/canonical/{canonical_digest.removeprefix('sha256:')}.json"
            ),
            content=canonical_content,
            media_type="application/vnd.knowledge.structured-dataset+json",
        )
        manifest_content = _canonical_json_bytes(
            {
                "schema_version": "dataset-record-manifest.v1",
                "knowledge_source_version_id": source_version_id,
                "dataset_revision_id": dataset_revision_id,
                "canonical_artifact_sha256": canonical.sha256,
                "records": [
                    {
                        "record_id": record.record_id,
                        "content_hash": record.content_hash,
                    }
                    for record in records
                ],
            }
        )
        evidence_manifest = self._put_and_verify(
            object_key=f"{key_root}/evidence-unit-manifest.json",
            content=manifest_content,
            media_type="application/vnd.knowledge.dataset-record-manifest+json",
        )
        publication = PublishedDatasetSourceVersion(
            version=version,
            original_artifact=original,
            canonical_artifact=canonical,
            evidence_manifest_artifact=evidence_manifest,
            processing_lineage_digest=processing_lineage_digest,
        )
        self._catalog.put_dataset_source_version(publication)
        return publication

    def _put_and_verify(
        self,
        *,
        object_key: str,
        content: bytes,
        media_type: str,
    ) -> ExactArtifactReference:
        reference = self._artifacts.put_immutable(
            object_key=object_key,
            content=content,
            media_type=media_type,
        )
        if self._artifacts.get_exact(reference) != content:
            raise ValueError("finalized immutable dataset artifact failed exact verification")
        return reference


@dataclass(frozen=True)
class XlsxDatasetIntakeCommand:
    knowledge_space_id: str
    knowledge_source_id: str
    display_filename: str
    content: bytes
    field_types: Mapping[str, StructuredValueType]


class XlsxDatasetIntakeApplication(_BoundedDatasetIntakeApplication):
    """Publish one formula-free worksheet as a typed immutable Dataset Revision."""

    def create_source_version(
        self,
        command: XlsxDatasetIntakeCommand,
    ) -> PublishedDatasetSourceVersion:
        _validate_authority_id(command.knowledge_space_id, "knowledge_space_id")
        _validate_authority_id(command.knowledge_source_id, "knowledge_source_id")
        if type(command.content) is not bytes or not command.content:
            raise ValueError("XLSX content must be nonempty exact bytes")
        if len(command.content) > self._max_content_bytes:
            raise ValueError("XLSX content exceeds the admitted size limit")
        if not command.content.startswith(b"PK\x03\x04"):
            raise ValueError("XLSX content signature is invalid")
        _validate_xlsx_package(command.content)
        field_types = tuple(command.field_types.items())
        if not field_types or any(not field.strip() for field, _ in field_types):
            raise ValueError("XLSX intake requires nonblank declared fields")
        if len({field for field, _ in field_types}) != len(field_types):
            raise ValueError("XLSX field declarations must be unique")
        workbook = load_workbook(
            BytesIO(command.content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        try:
            if len(workbook.worksheets) != 1:
                raise ValueError("XLSX intake requires exactly one worksheet")
            worksheet = workbook.worksheets[0]
            if worksheet.sheet_state != "visible":
                raise ValueError("XLSX worksheet must be visible")
            rows = worksheet.iter_rows()
            header = next(rows, None)
            field_order = tuple(field for field, _ in field_types)
            if header is None or tuple(cell.value for cell in header) != field_order:
                raise ValueError(
                    "XLSX headers must exactly match the declared field order"
                )
            fields_by_row: list[tuple[StructuredField, ...]] = []
            for row in rows:
                if len(row) != len(field_types):
                    raise ValueError("XLSX row width does not match the declared schema")
                if all(cell.value is None for cell in row):
                    continue
                if any(cell.data_type in {"f", "e"} for cell in row):
                    raise ValueError("XLSX formulas and error cells are forbidden")
                fields_by_row.append(
                    tuple(
                        _xlsx_field(
                            field=field,
                            declared_type=declared_type,
                            raw_value=cell.value,
                        )
                        for cell, (field, declared_type) in zip(
                            row,
                            field_types,
                            strict=True,
                        )
                    )
                )
                if len(fields_by_row) > self._max_records:
                    raise ValueError("XLSX dataset exceeds the admitted record limit")
            if not fields_by_row:
                raise ValueError("XLSX dataset must contain at least one record")
            worksheet_title = worksheet.title
        finally:
            workbook.close()
        return _publish_typed_dataset(
            artifacts=self._artifacts,
            catalog=self._catalog,
            pipeline_revision=self._pipeline_revision,
            knowledge_space_id=command.knowledge_space_id,
            knowledge_source_id=command.knowledge_source_id,
            content=command.content,
            media_type=_XLSX_MEDIA_TYPE,
            extension="xlsx",
            format_name="xlsx",
            lineage_fields={"worksheet": worksheet_title},
            field_types=field_types,
            fields_by_row=tuple(fields_by_row),
        )


@dataclass(frozen=True)
class ParquetDatasetIntakeCommand:
    knowledge_space_id: str
    knowledge_source_id: str
    display_filename: str
    content: bytes
    field_types: Mapping[str, StructuredValueType]


class ParquetDatasetIntakeApplication(_BoundedDatasetIntakeApplication):
    """Publish bounded primitive Parquet columns as a typed Dataset Revision."""

    def create_source_version(
        self,
        command: ParquetDatasetIntakeCommand,
    ) -> PublishedDatasetSourceVersion:
        _validate_authority_id(command.knowledge_space_id, "knowledge_space_id")
        _validate_authority_id(command.knowledge_source_id, "knowledge_source_id")
        if type(command.content) is not bytes or not command.content:
            raise ValueError("Parquet content must be nonempty exact bytes")
        if len(command.content) > self._max_content_bytes:
            raise ValueError("Parquet content exceeds the admitted size limit")
        if not (
            command.content.startswith(b"PAR1") and command.content.endswith(b"PAR1")
        ):
            raise ValueError("Parquet content signature is invalid")
        field_types = tuple(command.field_types.items())
        if not field_types or any(not field.strip() for field, _ in field_types):
            raise ValueError("Parquet intake requires nonblank declared fields")
        if len({field for field, _ in field_types}) != len(field_types):
            raise ValueError("Parquet field declarations must be unique")
        try:
            parquet = pq.ParquetFile(pa.BufferReader(command.content))
        except (pa.ArrowInvalid, OSError) as error:
            raise ValueError("Parquet structure is invalid") from error
        field_order = tuple(field for field, _ in field_types)
        schema = parquet.schema_arrow
        if tuple(schema.names) != field_order:
            raise ValueError(
                "Parquet columns must exactly match the declared field order"
            )
        if any(
            not _parquet_type_matches(schema.field(field).type, declared_type)
            for field, declared_type in field_types
        ):
            raise ValueError("Parquet physical type does not match declared schema")
        if (
            parquet.metadata.num_rows < 1
            or parquet.metadata.num_rows > self._max_records
            or parquet.metadata.num_row_groups < 1
            or parquet.metadata.num_row_groups > _MAX_PARQUET_ROW_GROUPS
        ):
            raise ValueError("Parquet dataset is outside the admitted bounds")
        try:
            rows = parquet.read(columns=list(field_order), use_threads=False).to_pylist()
        except (pa.ArrowInvalid, OSError) as error:
            raise ValueError("Parquet row decoding failed") from error
        if len(rows) != parquet.metadata.num_rows:
            raise ValueError("Parquet row count does not match its metadata")
        fields_by_row = tuple(
            tuple(
                _parquet_field(
                    field=field,
                    declared_type=declared_type,
                    raw_value=row[field],
                )
                for field, declared_type in field_types
            )
            for row in rows
        )
        return _publish_typed_dataset(
            artifacts=self._artifacts,
            catalog=self._catalog,
            pipeline_revision=self._pipeline_revision,
            knowledge_space_id=command.knowledge_space_id,
            knowledge_source_id=command.knowledge_source_id,
            content=command.content,
            media_type=_PARQUET_MEDIA_TYPE,
            extension="parquet",
            format_name="parquet",
            lineage_fields={
                "physical_schema": str(schema),
                "row_group_count": parquet.metadata.num_row_groups,
            },
            field_types=field_types,
            fields_by_row=fields_by_row,
        )


def _parquet_type_matches(
    physical_type: object,
    declared_type: StructuredValueType,
) -> bool:
    if declared_type == "string":
        return bool(
            pa.types.is_string(physical_type)
            or pa.types.is_large_string(physical_type)
        )
    if declared_type == "integer":
        return bool(pa.types.is_integer(physical_type))
    if declared_type == "decimal":
        return bool(pa.types.is_decimal(physical_type))
    if declared_type == "boolean":
        return bool(pa.types.is_boolean(physical_type))
    if declared_type == "date":
        return bool(pa.types.is_date(physical_type))
    if declared_type == "datetime":
        return bool(
            pa.types.is_timestamp(physical_type)
            and getattr(physical_type, "tz", None) is not None
        )
    return bool(pa.types.is_null(physical_type))


def _parquet_field(
    *,
    field: str,
    declared_type: StructuredValueType,
    raw_value: object,
) -> StructuredField:
    if raw_value is None:
        return StructuredField(field=field, value_type="null", value=None)
    if declared_type == "string" and type(raw_value) is str:
        return StructuredField(field=field, value_type="string", value=raw_value)
    if declared_type == "integer" and type(raw_value) is int:
        return StructuredField(field=field, value_type="integer", value=raw_value)
    if declared_type == "decimal" and isinstance(raw_value, Decimal):
        if raw_value.is_finite():
            return StructuredField(
                field=field,
                value_type="decimal",
                value=format(raw_value, "f"),
            )
    if declared_type == "boolean" and type(raw_value) is bool:
        return StructuredField(field=field, value_type="boolean", value=raw_value)
    if declared_type == "date" and type(raw_value) is date:
        return StructuredField(
            field=field,
            value_type="date",
            value=raw_value.isoformat(),
        )
    if declared_type == "datetime" and isinstance(raw_value, datetime):
        if raw_value.tzinfo is not None and raw_value.utcoffset() is not None:
            return StructuredField(
                field=field,
                value_type="datetime",
                value=raw_value.isoformat(),
            )
    raise ValueError(f"Parquet value does not match declared type for field {field}")


def _validate_xlsx_package(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            infos = archive.infolist()
            names = [info.filename for info in infos]
            lowered = {name.casefold() for name in names}
            if (
                not infos
                or len(infos) > _MAX_XLSX_PARTS
                or len(set(names)) != len(names)
                or any(
                    info.flag_bits & 1
                    or info.filename.startswith(("/", "\\"))
                    or ".." in info.filename.replace("\\", "/").split("/")
                    for info in infos
                )
                or sum(info.file_size for info in infos)
                > _MAX_XLSX_UNCOMPRESSED_BYTES
                or "[content_types].xml" not in lowered
                or "xl/workbook.xml" not in lowered
                or any(
                    name.endswith("vbaproject.bin")
                    or name.startswith("xl/externallinks/")
                    or "/embeddings/" in name
                    for name in lowered
                )
            ):
                raise ValueError("XLSX package violates the admitted safety profile")
    except BadZipFile as error:
        raise ValueError("XLSX package is invalid") from error


def _xlsx_field(
    *,
    field: str,
    declared_type: StructuredValueType,
    raw_value: object,
) -> StructuredField:
    if raw_value is None:
        return StructuredField(field=field, value_type="null", value=None)
    if declared_type == "string" and type(raw_value) is str:
        return StructuredField(field=field, value_type="string", value=raw_value)
    if declared_type == "integer" and type(raw_value) is int:
        return StructuredField(field=field, value_type="integer", value=raw_value)
    if declared_type == "decimal" and type(raw_value) in {int, float, Decimal}:
        decimal_value = Decimal(str(raw_value))
        if decimal_value.is_finite():
            return StructuredField(
                field=field,
                value_type="decimal",
                value=format(decimal_value, "f"),
            )
    if declared_type == "boolean" and type(raw_value) is bool:
        return StructuredField(field=field, value_type="boolean", value=raw_value)
    if declared_type in {"date", "datetime"} and type(raw_value) is str:
        return _parse_field(
            field=field,
            declared_type=declared_type,
            raw_value=raw_value,
        )
    raise ValueError(f"XLSX value does not match declared type for field {field}")


def _publish_typed_dataset(
    *,
    artifacts: ImmutableArtifactStore,
    catalog: KnowledgeCatalogWriter,
    pipeline_revision: str,
    knowledge_space_id: str,
    knowledge_source_id: str,
    content: bytes,
    media_type: str,
    extension: str,
    format_name: str,
    lineage_fields: Mapping[str, object],
    field_types: tuple[tuple[str, StructuredValueType], ...],
    fields_by_row: tuple[tuple[StructuredField, ...], ...],
) -> PublishedDatasetSourceVersion:
    original_digest = f"sha256:{sha256(content).hexdigest()}"
    schema_revision_id = content_identifier(
        "dataset-schema",
        sha256_json({"fields": field_types}),
    )
    processing_lineage_digest = sha256_json(
        {
            "schema": "dataset-processing-lineage.v1",
            "pipeline_revision": pipeline_revision,
            "format": format_name,
            "schema_revision_id": schema_revision_id,
            **lineage_fields,
        }
    )
    version_digest = sha256_json(
        {
            "knowledge_space_id": knowledge_space_id,
            "knowledge_source_id": knowledge_source_id,
            "original_digest": original_digest,
            "schema_revision_id": schema_revision_id,
            "processing_lineage_digest": processing_lineage_digest,
        }
    )
    source_version_id = content_identifier("source-version", version_digest)
    dataset_revision_id = content_identifier("dataset-revision", version_digest)
    records = tuple(
        _record_from_fields(
            row_number=row_number,
            fields=fields,
            dataset_revision_id=dataset_revision_id,
        )
        for row_number, fields in enumerate(fields_by_row, start=1)
    )
    version = DatasetSourceVersion(
        knowledge_space_id=knowledge_space_id,
        knowledge_source_id=knowledge_source_id,
        knowledge_source_version_id=source_version_id,
        dataset_revision_id=dataset_revision_id,
        schema_revision_id=schema_revision_id,
        field_order=tuple(field for field, _ in field_types),
        records=records,
    )
    key_root = (
        f"spaces/{knowledge_space_id}/sources/{knowledge_source_id}/"
        f"versions/{source_version_id}"
    )
    original = _put_exact(
        artifacts=artifacts,
        object_key=(
            f"{key_root}/originals/{original_digest.removeprefix('sha256:')}."
            f"{extension}"
        ),
        content=content,
        media_type=media_type,
    )
    canonical_content = _canonical_json_bytes(
        {
            "schema_version": "structured-dataset-revision.v1",
            "knowledge_source_version_id": source_version_id,
            "dataset_revision_id": dataset_revision_id,
            "schema_revision_id": schema_revision_id,
            "processing_lineage_digest": processing_lineage_digest,
            "fields": [
                {"field": field, "value_type": value_type}
                for field, value_type in field_types
            ],
            "records": [_record_payload(record) for record in records],
        }
    )
    canonical_digest = f"sha256:{sha256(canonical_content).hexdigest()}"
    canonical = _put_exact(
        artifacts=artifacts,
        object_key=(
            f"{key_root}/canonical/{canonical_digest.removeprefix('sha256:')}.json"
        ),
        content=canonical_content,
        media_type="application/vnd.knowledge.structured-dataset+json",
    )
    manifest_content = _canonical_json_bytes(
        {
            "schema_version": "dataset-record-manifest.v1",
            "knowledge_source_version_id": source_version_id,
            "dataset_revision_id": dataset_revision_id,
            "canonical_artifact_sha256": canonical.sha256,
            "records": [
                {
                    "record_id": record.record_id,
                    "content_hash": record.content_hash,
                }
                for record in records
            ],
        }
    )
    evidence_manifest = _put_exact(
        artifacts=artifacts,
        object_key=f"{key_root}/evidence-unit-manifest.json",
        content=manifest_content,
        media_type="application/vnd.knowledge.dataset-record-manifest+json",
    )
    publication = PublishedDatasetSourceVersion(
        version=version,
        original_artifact=original,
        canonical_artifact=canonical,
        evidence_manifest_artifact=evidence_manifest,
        processing_lineage_digest=processing_lineage_digest,
    )
    catalog.put_dataset_source_version(publication)
    return publication


def _put_exact(
    *,
    artifacts: ImmutableArtifactStore,
    object_key: str,
    content: bytes,
    media_type: str,
) -> ExactArtifactReference:
    reference = artifacts.put_immutable(
        object_key=object_key,
        content=content,
        media_type=media_type,
    )
    if artifacts.get_exact(reference) != content:
        raise ValueError("finalized immutable dataset artifact failed exact verification")
    return reference


def _record_payload(record: StructuredRecord) -> dict[str, object]:
    return {
        "record_id": record.record_id,
        "content_hash": record.content_hash,
        "fields": [
            {
                "field": field.field,
                "value_type": field.value_type,
                "value": field.value,
            }
            for field in record.fields
        ],
    }


def _record_from_fields(
    *,
    row_number: int,
    fields: tuple[StructuredField, ...],
    dataset_revision_id: str,
) -> StructuredRecord:
    record_payload = [
        {"field": item.field, "value_type": item.value_type, "value": item.value}
        for item in fields
    ]
    content_hash = sha256_json(record_payload)
    record_id = content_identifier(
        "record",
        sha256_json(
            {
                "dataset_revision": dataset_revision_id,
                "row_number": row_number,
                "content_hash": content_hash,
            }
        ),
    )
    return StructuredRecord(record_id=record_id, fields=fields, content_hash=content_hash)


def _record_from_row(
    *,
    row_number: int,
    row: Mapping[str | None, str | list[str] | None],
    field_types: tuple[tuple[str, StructuredValueType], ...],
    dataset_revision_id: str,
) -> StructuredRecord:
    if None in row:
        raise ValueError("CSV record has more values than declared fields")
    fields = tuple(
        _parse_field(
            field=field,
            declared_type=value_type,
            raw_value=_raw_csv_value(row.get(field), field),
        )
        for field, value_type in field_types
    )
    record_payload = [
        {"field": item.field, "value_type": item.value_type, "value": item.value}
        for item in fields
    ]
    content_hash = sha256_json(record_payload)
    record_id = content_identifier(
        "record",
        sha256_json(
            {
                "dataset_revision": dataset_revision_id,
                "row_number": row_number,
                "content_hash": content_hash,
            }
        ),
    )
    return StructuredRecord(record_id=record_id, fields=fields, content_hash=content_hash)


def _raw_csv_value(value: str | list[str] | None, field: str) -> str | None:
    if isinstance(value, list):
        raise ValueError(f"CSV field {field} contains an invalid repeated value")
    return value


def _parse_field(
    *,
    field: str,
    declared_type: StructuredValueType,
    raw_value: str | None,
) -> StructuredField:
    if raw_value is None or raw_value == "":
        return StructuredField(field=field, value_type="null", value=None)
    value: StructuredScalar
    if declared_type == "string":
        value = raw_value
    elif declared_type == "integer":
        value = int(raw_value)
    elif declared_type == "decimal":
        try:
            decimal_value = Decimal(raw_value)
        except InvalidOperation as error:
            raise ValueError(f"invalid decimal in field {field}") from error
        if not decimal_value.is_finite():
            raise ValueError(f"non-finite decimal in field {field}")
        value = format(decimal_value, "f")
    elif declared_type == "boolean":
        normalized = raw_value.casefold()
        if normalized not in {"true", "false"}:
            raise ValueError(f"invalid boolean in field {field}")
        value = normalized == "true"
    elif declared_type == "date":
        value = date.fromisoformat(raw_value).isoformat()
    elif declared_type == "datetime":
        parsed = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
        if parsed.tzinfo is None or parsed.utcoffset() is None:
            raise ValueError(f"naive datetime in field {field}")
        value = parsed.isoformat()
    elif declared_type == "null":
        raise ValueError(f"non-null value in null-only field {field}")
    else:
        raise ValueError(f"unsupported structured type for field {field}")
    return StructuredField(field=field, value_type=declared_type, value=value)


def _canonical_json_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")


def _validate_authority_id(value: str, field: str) -> None:
    if _AUTHORITY_ID.fullmatch(value) is None:
        raise ValueError(f"{field} is not a valid opaque authority identifier")
